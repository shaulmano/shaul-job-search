import sys
sys.stdout.reconfigure(encoding='utf-8')

import csv, time, random, os, re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

RECRUITERS_FILE = r"C:\Users\Shaul\Documents\job-search\phantom_recruiters.csv"
PROGRESS_FILE   = r"C:\Users\Shaul\Documents\job-search\harvest_done.txt"
OUTREACH_FILE   = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"


def load_done():
    if not os.path.exists(PROGRESS_FILE):
        return set()
    with open(PROGRESS_FILE, encoding='utf-8') as f:
        return set(l.strip() for l in f if l.strip())


def mark_done(url):
    with open(PROGRESS_FILE, 'a', encoding='utf-8') as f:
        f.write(url + '\n')


def load_existing_emails():
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active
    return set(
        str(r[3].value or '').strip().lower()
        for r in ws.iter_rows(min_row=2) if r[3].value
    )


def ensure_source_column(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if 'Source' not in headers:
        ws.cell(1, ws.max_column + 1).value = 'Source'
    return ws.max_column


def save_email(first, last, company, email, linkedin_url):
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active
    src_col = ensure_source_column(ws)
    row = [first, last, company, email, linkedin_url, 'Pending', '']
    # השלם עמודות ריקות עד Source
    while len(row) < src_col - 1:
        row.append('')
    row.append('LinkedIn_Harvest')
    ws.append(row)
    wb.save(OUTREACH_FILE)


def extract_email_from_text(text):
    m = re.search(r'[\w.%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}', text)
    return m.group(0).lower() if m else None


def init_browser():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    return webdriver.Chrome(options=options)


def _js_set_value(driver, css_selector, value):
    driver.execute_script("""
        var field = document.querySelector(arguments[0]);
        var nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        nativeSetter.call(field, arguments[1]);
        field.dispatchEvent(new Event('input',  { bubbles: true }));
        field.dispatchEvent(new Event('change', { bubbles: true }));
    """, css_selector, value)


def login(driver):
    driver.get('https://www.linkedin.com/login')
    wait = WebDriverWait(driver, 20)

    print("Waiting for login page to load...")
    time.sleep(5)
    print(f"  Page title: {driver.title}")
    print(f"  Current URL: {driver.current_url}")

    # dismiss cookie banner if present
    for dismiss_sel in [
        "//button[contains(@action-type,'DENY')]",
        "//button[contains(text(),'Reject')]",
        "//button[contains(text(),'Accept')]",
    ]:
        try:
            driver.find_element(By.XPATH, dismiss_sel).click()
            time.sleep(1)
            break
        except:
            pass

    print("  Filling email via JS...")
    wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='email' and @autocomplete='username webauthn']")))
    _js_set_value(driver, "input[autocomplete='username webauthn']", LINKEDIN_EMAIL)
    print(f"  Email set: {LINKEDIN_EMAIL}")

    time.sleep(0.8)

    print("  Filling password via JS...")
    # pick the visible password field (second one)
    pwd_fields = driver.find_elements(By.XPATH, "//input[@type='password']")
    visible_pwd = next((f for f in pwd_fields if f.is_displayed()), pwd_fields[-1])
    driver.execute_script("""
        var field = arguments[0];
        var nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        nativeSetter.call(field, arguments[1]);
        field.dispatchEvent(new Event('input',  { bubbles: true }));
        field.dispatchEvent(new Event('change', { bubbles: true }));
    """, visible_pwd, LINKEDIN_PASSWORD)
    print("  Password set.")

    time.sleep(0.8)

    # click submit via JavaScript to bypass interactability issues
    clicked = driver.execute_script("""
        var btn = document.querySelector('button[type="submit"]');
        if (!btn) btn = document.querySelector('button[data-litms-control-urn]');
        if (!btn) {
            var btns = document.querySelectorAll('button');
            for (var b of btns) { if (b.textContent.trim().toLowerCase().includes('sign in')) { btn = b; break; } }
        }
        if (btn) { btn.click(); return true; }
        return false;
    """)
    print(f"  Submit clicked via JS: {clicked}")
    time.sleep(6)

    if 'checkpoint' in driver.current_url or 'challenge' in driver.current_url:
        print("LinkedIn verification required — complete it in the browser and press Enter.")
        input(">> Enter: ")

    print(f"  Current URL: {driver.current_url}")
    print("Logged in.\n")


def get_contact_email(driver, profile_url, wait):
    driver.get(profile_url)
    time.sleep(random.uniform(3, 5))

    try:
        contact_link = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//a[contains(@href,'detail/contact-info')]")
        ))
        contact_link.click()
        time.sleep(3)

        try:
            # wait for any dialog/section that contains an email link
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//a[starts-with(@href,'mailto:')]"))
            )
            mailto_links = driver.find_elements(By.XPATH, "//a[starts-with(@href,'mailto:')]")
            if mailto_links:
                email = mailto_links[0].get_attribute('href').replace('mailto:', '').strip().lower()
                try:
                    driver.find_element(By.XPATH, "//button[@aria-label='Dismiss']").click()
                except:
                    pass
                return email
        except:
            pass

        # fallback: search full page text for email pattern
        try:
            page_text = driver.find_element(By.TAG_NAME, 'body').text
            email = extract_email_from_text(page_text)
            try:
                driver.find_element(By.XPATH, "//button[@aria-label='Dismiss']").click()
            except:
                pass
            return email
        except:
            pass

        return None

    except Exception:
        return None


def main():
    with open(RECRUITERS_FILE, encoding='utf-8-sig') as f:
        recruiters = list(csv.DictReader(f))

    done     = load_done()
    existing = load_existing_emails()
    pending  = [r for r in recruiters if r['linkedinProfileUrl'].strip() not in done]

    test_mode = '--test' in sys.argv
    if test_mode:
        pending = pending[:3]
        print("*** TEST MODE — 3 profiles only ***\n")

    print(f"Total profiles : {len(recruiters)}")
    print(f"Already done   : {len(done)}")
    print(f"Remaining      : {len(pending)}")
    if not test_mode:
        print(f"Estimated time : {len(pending) * 30 // 60} min ({len(pending) * 30 / 3600:.1f} hours)")
    print()
    input("Press Enter to start (make sure Chrome is open and you're logged into LinkedIn)...")

    driver = init_browser()
    wait   = WebDriverWait(driver, 8)
    print(f"Connected to browser. URL: {driver.current_url}\n")

    found = 0
    start = datetime.now()

    try:
        for i, r in enumerate(pending, 1):
            url   = r['linkedinProfileUrl'].strip()
            first = r['firstName'].strip()
            last  = r['lastName'].strip()

            elapsed = (datetime.now() - start).seconds // 60
            print(f"[{i}/{len(pending)}] {elapsed}m | {first} {last} ... ", end='', flush=True)

            email = get_contact_email(driver, url, wait)
            mark_done(url)

            if email and email not in existing and '@' in email:
                save_email(first, last, '', email, url)
                existing.add(email)
                found += 1
                print(f"Found: {email}")
            else:
                print("No email")

            time.sleep(random.uniform(20, 35))

    except KeyboardInterrupt:
        print("\nStopped by user.")

    finally:
        print(f"\n=== DONE ===")
        print(f"New emails found: {found}")
        input("Press Enter to close browser...")
        driver.quit()


if __name__ == '__main__':
    main()
