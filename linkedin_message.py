import sys
sys.stdout.reconfigure(encoding='utf-8')

import csv
import time
import random
import os
import re
import openpyxl
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from config import LINKEDIN_EMAIL, LINKEDIN_PASSWORD, SENDER_PHONE, SENDER_LINKEDIN

PROGRESS_FILE = r"C:\Users\Shaul\Documents\job-search\linkedin_progress.csv"
XLSX          = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"
DAILY_LIMIT   = 20
PREVIEW_COUNT = 5

PDF_1 = r"C:\Users\Shaul\Documents\job-search\Shaul_Impact.pdf"
PDF_2 = r"C:\Users\Shaul\Documents\job-search\Shaul_mano_Project.pdf"
PDF_3 = r"C:\Users\Shaul\Documents\job-search\Shaul_mano_QA.pdf"

FULL_MESSAGE = """\
Hi {first_name},

Thanks for connecting!

I'm currently looking for my next role as a Senior Program Manager or QA Leader, with 25+ years of experience at companies like RSA and Symantec.
I'd love to hear if you have anything relevant open. I've attached my CVs below — feel free to share whichever fits best:

Impact Summary: https://drive.google.com/file/d/1h3PObY_Ap_6HLtNB0AApvWzR1rUu-qii/view?usp=sharing
Program Management: https://drive.google.com/file/d/1B3SmlKLlDuCRVxC_jG7f4WeMJnTW4cUK/view?usp=sharing
QA Leadership: https://drive.google.com/file/d/1HDIF60dWsUFVc3lBpsH3l--Bjy46We5t/view?usp=sharing

Happy to jump on a quick 15-minute call if anything looks like a fit.

Best,
Shaul Mano
{phone}
{linkedin}
"""


def load_progress():
    if not os.path.exists(PROGRESS_FILE):
        return []
    with open(PROGRESS_FILE, encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


def slug(url):
    m = re.search(r'linkedin\.com/in/([^/?#]+)', str(url or '').lower())
    return m.group(1).rstrip('/') if m else None

def update_status(url, new_status):
    # Update CSV progress file
    rows = load_progress()
    with open(PROGRESS_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['firstName', 'lastName', 'linkedinProfileUrl', 'status', 'date'])
        writer.writeheader()
        for row in rows:
            if row['linkedinProfileUrl'].strip() == url:
                row['status'] = new_status
                row['date'] = str(date.today())
            writer.writerow(row)
    # Update master xlsx
    if new_status == 'Message_Sent' and os.path.exists(XLSX):
        try:
            wb = openpyxl.load_workbook(XLSX)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            li_col = headers.index('LinkedIn URL') + 1
            st_col = headers.index('Status') + 1
            dt_col = headers.index('Date Sent') + 1
            s = slug(url)
            for i in range(2, ws.max_row + 1):
                if slug(ws.cell(i, li_col).value) == s:
                    ws.cell(i, st_col).value = 'Message Sent'
                    ws.cell(i, dt_col).value = str(date.today())
                    break
            wb.save(XLSX)
        except Exception as e:
            print(f"    (xlsx update failed: {e})")


def pause(min_sec=2, max_sec=5):
    time.sleep(random.uniform(min_sec, max_sec))


def init_browser():
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver):
    driver.get('https://www.linkedin.com/login')
    wait = WebDriverWait(driver, 15)
    wait.until(EC.presence_of_element_located((By.ID, 'username'))).send_keys(LINKEDIN_EMAIL)
    driver.find_element(By.ID, 'password').send_keys(LINKEDIN_PASSWORD)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    pause(4, 7)

    if 'checkpoint' in driver.current_url or 'challenge' in driver.current_url:
        print("\n  LinkedIn מבקש אימות — השלם אותו בדפדפן ואז לחץ Enter כאן.")
        input("  >> לחץ Enter לאחר האימות: ")
        pause(2, 3)
    print("  Connected to LinkedIn.\n")


def is_connected(driver, url):
    """Returns True if the profile shows a Message button (= accepted connection)."""
    driver.get(url)
    pause(3, 5)
    try:
        driver.find_element(
            By.XPATH,
            "//button[.//span[text()='Message']]|//a[.//span[text()='Message']]"
        )
        return True
    except Exception:
        return False


def send_message_with_attachments(driver, url, first_name):
    wait = WebDriverWait(driver, 10)

    try:
        # Click Message button
        msg_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[text()='Message']]|//a[.//span[text()='Message']]")
        ))
        msg_btn.click()
        pause(2, 3)

        # Type the message
        msg_box = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//div[@role='textbox' and @contenteditable='true']")
        ))
        msg_box.click()
        full_msg = FULL_MESSAGE.format(
            first_name=first_name,
            phone=SENDER_PHONE,
            linkedin=SENDER_LINKEDIN
        )
        msg_box.send_keys(full_msg)
        pause(1, 2)

        # Attach files
        for pdf_path in [PDF_1, PDF_2, PDF_3]:
            if not os.path.exists(pdf_path):
                print(f"\n    WARNING: file not found: {pdf_path}")
                continue
            try:
                attach_input = driver.find_element(
                    By.XPATH, "//input[@type='file']"
                )
                attach_input.send_keys(pdf_path)
                pause(2, 3)
                print(f"    Attached: {os.path.basename(pdf_path)}")
            except Exception as e:
                print(f"\n    Could not attach {os.path.basename(pdf_path)}: {e}")

        pause(1, 2)

        # Send
        send_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(@class,'msg-form__send-button')]"
                       "|//button[@type='submit' and .//span[text()='Send']]"
                       "|//button[contains(@aria-label,'Send')]")
        ))
        send_btn.click()
        pause(2, 3)
        return 'sent'

    except Exception as e:
        return f'error: {e}'


def main():
    all_rows = load_progress()

    # Only process those who got a connection request but haven't received the full message yet
    to_message = [
        r for r in all_rows
        if r['status'].strip() == 'Connect_Sent'
    ]

    print(f"Connection requests sent : {len([r for r in all_rows if r['status'] == 'Connect_Sent'])}")
    print(f"Already messaged         : {len([r for r in all_rows if r['status'] == 'Message_Sent'])}")
    print(f"To check for acceptance  : {len(to_message)}")

    if not to_message:
        print("\nNo pending profiles to message.")
        return

    batch = to_message[:DAILY_LIMIT]
    full_msg_sample = FULL_MESSAGE.format(
        first_name=batch[0]['firstName'].strip(),
        phone=SENDER_PHONE,
        linkedin=SENDER_LINKEDIN
    )

    print(f"\n{'='*60}")
    print(f"PREVIEW — checking up to {len(batch)} profiles for accepted connections")
    print(f"{'='*60}")
    for i, r in enumerate(batch[:PREVIEW_COUNT], 1):
        print(f"\n[{i}] {r['firstName'].strip()} {r['lastName'].strip()}")
        print(f"    {r['linkedinProfileUrl']}")
    print(f"\nMessage to send (with 3 PDF attachments):")
    print(f"{'-'*40}")
    print(full_msg_sample)
    print(f"Attachments: Shaul_Impact.pdf | Shaul_mano_Project.pdf | Shaul_mano_QA.pdf")
    print(f"{'='*60}")

    answer = input(f"\nCheck {len(batch)} profiles and send full message to those who accepted? (yes / no): ").strip().lower()
    if answer != 'yes':
        print("Cancelled.")
        return

    print("\nOpening browser...")
    driver = init_browser()

    try:
        login(driver)

        checked = 0
        sent_count = 0
        not_yet = 0

        for i, r in enumerate(batch, 1):
            first = r['firstName'].strip()
            last  = r['lastName'].strip()
            url   = r['linkedinProfileUrl'].strip()

            print(f"[{i}/{len(batch)}] {first} {last} ... ", end='', flush=True)
            checked += 1

            if is_connected(driver, url):
                print("Connected! Sending message... ", end='', flush=True)
                result = send_message_with_attachments(driver, url, first)
                if result == 'sent':
                    update_status(url, 'Message_Sent')
                    sent_count += 1
                    print("Done")
                else:
                    print(f"Error — {result}")
            else:
                not_yet += 1
                print("Not yet accepted")

            pause(4, 7)

    finally:
        print(f"\n{'='*40}")
        print(f"Checked  : {checked}")
        print(f"Messaged : {sent_count}")
        print(f"Pending  : {not_yet}")
        input("\nPress Enter to close browser...")
        driver.quit()


if __name__ == '__main__':
    main()
