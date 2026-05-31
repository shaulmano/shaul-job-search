import sys
sys.stdout.reconfigure(encoding='utf-8')

import imaplib
import email
import re
from datetime import date, timedelta
from openpyxl import load_workbook
from config import GMAIL_ADDRESS, GMAIL_APP_PASSWORD, OUTREACH_FILE


def connect_gmail():
    mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
    mail.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
    return mail


def extract_bounced_address(msg):
    """Extract the original To address from a bounce message."""
    # Method 1: DSN format — "Final-Recipient: rfc822; email"
    for part in msg.walk():
        ct = part.get_content_type()
        if ct in ("message/delivery-status", "text/plain"):
            try:
                payload = part.get_payload(decode=True)
                if payload:
                    text = payload.decode(errors="replace")
                else:
                    # delivery-status parts may be non-encoded sub-messages
                    text = str(part.get_payload())
                m = re.search(r'Final-Recipient\s*:\s*rfc822\s*;\s*([^\s\r\n]+)', text, re.IGNORECASE)
                if m:
                    return m.group(1).strip().lower()
                m = re.search(r'Original-Recipient\s*:\s*rfc822\s*;\s*([^\s\r\n]+)', text, re.IGNORECASE)
                if m:
                    return m.group(1).strip().lower()
            except Exception:
                pass

    # Method 2: attached original message — look at its To header
    for part in msg.walk():
        if part.get_content_type() == "message/rfc822":
            try:
                orig = part.get_payload(0)
                to_header = orig.get("To", "")
                m = re.search(r'[\w.%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}', to_header)
                if m:
                    return m.group(0).strip().lower()
            except Exception:
                pass

    # Method 3: scan body text for "to <email>" pattern
    for part in msg.walk():
        if part.get_content_type() == "text/plain":
            try:
                text = part.get_payload(decode=True).decode(errors="replace")
                # look for lines like: "Your message to X failed" or "couldn't deliver to X"
                for pattern in [
                    r'(?:to|for|recipient|address)\s*[:<]?\s*([\w.%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})',
                    r'([\w.%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})',
                ]:
                    candidates = re.findall(pattern, text, re.IGNORECASE)
                    for c in candidates:
                        c = c.lower()
                        if c != GMAIL_ADDRESS.lower() and "mailer-daemon" not in c:
                            return c
            except Exception:
                pass

    return None


def fetch_bounces(days_back=14):
    """Connect to Gmail and return set of bounced email addresses."""
    mail = connect_gmail()
    bounced = set()

    since_date = (date.today() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    search_query = f'(FROM "mailer-daemon" SINCE {since_date})'

    for folder in ['INBOX', '[Gmail]/Trash', '[Gmail]/All Mail']:
        try:
            mail.select(folder, readonly=True)
            _, data = mail.search(None, search_query)
            ids = data[0].split()
            if not ids or ids == [b'']:
                continue
            print(f"  {folder}: {len(ids)} bounce message(s)")
            for uid in ids:
                _, msg_data = mail.fetch(uid, "(RFC822)")
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                addr = extract_bounced_address(msg)
                if addr and re.match(r'^[\w.%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}$', addr):
                    bounced.add(addr)
        except Exception as e:
            print(f"  Warning: could not search {folder}: {e}")

    mail.logout()
    return bounced


def mark_bounced_in_excel(bounced_emails):
    """Update outreach_list.xlsx: mark matching rows as Bounced."""
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    updated = 0
    for row in ws.iter_rows(min_row=2):
        email_cell = str(row[3].value or "").strip().lower()
        status_cell = str(row[5].value or "").strip()
        if email_cell in bounced_emails and status_cell != "Bounced":
            row[5].value = "Bounced"
            updated += 1
            print(f"  Marked Bounced: {email_cell}")

    wb.save(OUTREACH_FILE)
    return updated


def run_retry(bounced_emails):
    """Add next-pattern rows for all bounced emails."""
    from retry_bounced import next_pattern, normalize, get_domain
    import re as _re

    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    existing_emails = set()
    bounced_rows = []

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        em = str(row[3].value or "").strip().lower()
        status = str(row[5].value or "").strip()
        if em:
            existing_emails.add(em)
        if status == "Bounced" and em in bounced_emails:
            bounced_rows.append({
                "row": i,
                "first": str(row[0].value or "").strip(),
                "last":  str(row[1].value or "").strip(),
                "company": str(row[2].value or "").strip(),
                "email": em,
                "linkedin": str(row[4].value or "").strip(),
            })

    added = 0
    for item in bounced_rows:
        new_email = next_pattern(item["first"], item["last"], item["email"])
        if not new_email:
            print(f"  {item['email']} — no more patterns")
            continue
        if not _re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', new_email):
            print(f"  {item['email']} → {new_email} — invalid format, skipping")
            continue
        if new_email.lower() in existing_emails:
            print(f"  {item['email']} → {new_email} — already in list")
            continue
        ws.append([item["first"], item["last"], item["company"],
                   new_email, item["linkedin"], "Pending", ""])
        existing_emails.add(new_email.lower())
        added += 1
        print(f"  {item['email']}")
        print(f"  → {new_email} (added as Pending)")

    wb.save(OUTREACH_FILE)
    return added


def main():
    print(f"Checking Gmail for bounces (last 14 days)...\n")
    bounced = fetch_bounces(days_back=14)

    if not bounced:
        print("No new bounces found.")
        return

    print(f"\nFound {len(bounced)} bounced address(es):")
    for b in sorted(bounced):
        print(f"  {b}")

    print(f"\nUpdating outreach_list.xlsx...")
    updated = mark_bounced_in_excel(bounced)
    print(f"Marked {updated} new rows as Bounced.")

    if updated > 0:
        print(f"\nRunning retry logic...")
        added = run_retry(bounced)
        print(f"Added {added} new retry row(s).")

    print("\nDone.")


if __name__ == "__main__":
    main()
