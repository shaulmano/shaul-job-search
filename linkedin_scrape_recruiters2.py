import sys
sys.stdout.reconfigure(encoding='utf-8')

import csv, time, random, os, re
from urllib.parse import quote
from selenium import webdriver
import openpyxl

OUTPUT_FILE = r"C:\Users\Shaul\Documents\job-search\recruiters_scraped.csv"
DONE_FILE   = r"C:\Users\Shaul\Documents\job-search\recruiters_scraped_done.txt"

ISRAEL_GEO = "101620260"

KEYWORDS = [
    "HR Director",
    "Director of Talent Acquisition",
    "Head of Talent",
    "Head of HR",
    "Head of People",
    "VP HR",
    "VP People",
    "Chief People Officer",
    "Chief HR Officer",
    "People Director",
    "Director of People",
    "Talent Acquisition Manager",
    "Recruitment Manager",
    "HR Manager hi-tech",
    "People Manager",
    "Director of Human Resources",
]

MAX_PAGES_PER_KEYWORD = 10
MAX_TOTAL = 200

TECH_TITLE_KW = [
    'tech', 'technical', 'hi-tech', 'high-tech', 'hitech',
    'software', 'r&d', 'cyber', 'data scientist',
    ' ai', 'ai ', 'cloud', 'saas', 'fintech',
    'medtech', 'biotech', 'it recruiter',
    'startup', 'dev',
    'hr director', 'head of hr', 'head of talent', 'head of people',
    'vp hr', 'vp people', 'chief people', 'chief hr',
    'people director', 'director of people', 'director of talent',
    'talent acquisition manager', 'recruitment manager', 'people manager',
    'director of human', 'hr manager',
]

TECH_COMPANY_KW = [
    'tech', 'software', 'systems', 'solutions', 'digital',
    'cyber', 'cloud', 'networks', 'labs', 'technologies',
    'microsoft', 'google', 'amazon', 'apple', 'intel', 'nvidia',
    'cisco', 'ibm', 'salesforce', 'oracle', 'sap',
    'amdocs', 'check point', 'nice', 'wix', 'fiverr',
    'monday.com', 'mobileye', 'mellanox', 'radware', 'imperva',
    'varonis', 'cyberark', 'armis', 'sisense', 'riskified',
    'payoneer', 'taboola', 'outbrain', 'appsflyer', 'experis',
    'elbit', 'rafael', 'cellebrite', 'allot', 'nova', 'camtek',
    'matrix', 'sqlink', 'malam', 'aman', 'gotfriends', 'ness ',
    'bynet', 'sap', 'priority', 'lightricks', 'lemonade', 'elementor',
    'tipalti', 'ironscales', 'servicenow', 'workday', 'zendesk',
    'freshworks', 'hubspot', 'zero networks', 'bright data',
    'similarweb', 'similar web', 'ironwood', 'checkpoint',
    'qualcomm', 'broadcom', 'marvell', 'cadence', 'synopsys',
    'startup', 'innovation', 'ventures', 'r&d', 'devops',
    'infosec', 'deep learning', 'machine learning',
]

def is_tech(row):
    title = (row.get('jobTitle') or '').lower()
    company = (row.get('company') or '').lower()
    if any(kw in title for kw in TECH_TITLE_KW):
        return True
    if any(kw in company for kw in TECH_COMPANY_KW):
        return True
    return False

def connect():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    return webdriver.Chrome(options=options)

def load_done():
    if not os.path.exists(DONE_FILE):
        return set()
    with open(DONE_FILE, encoding='utf-8') as f:
        return set(line.strip() for line in f if line.strip())

def save_done(keyword):
    with open(DONE_FILE, 'a', encoding='utf-8') as f:
        f.write(keyword + '\n')

XLSX = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"

def slug(url):
    m = re.search(r'linkedin\.com/in/([^/?#]+)', str(url or '').lower())
    return m.group(1).rstrip('/') if m else None

def load_existing_urls():
    urls = set()
    # from CSV
    if os.path.exists(OUTPUT_FILE):
        with open(OUTPUT_FILE, encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                u = row.get('linkedinProfileUrl', '')
                if u:
                    urls.add(u)
                    s = slug(u)
                    if s: urls.add(s)
    # from outreach_list.xlsx
    if os.path.exists(XLSX):
        wb = openpyxl.load_workbook(XLSX, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.rows)]
        if 'LinkedIn URL' in headers:
            li_col = headers.index('LinkedIn URL')
            for row in ws.iter_rows(min_row=2, values_only=True):
                u = row[li_col]
                if u:
                    urls.add(str(u))
                    s = slug(str(u))
                    if s: urls.add(s)
        wb.close()
    print(f"Existing profiles (CSV + xlsx): {len(urls)}")
    return urls

def append_rows(rows):
    file_exists = os.path.exists(OUTPUT_FILE)
    with open(OUTPUT_FILE, 'a', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['firstName','lastName','linkedinProfileUrl','jobTitle','company'])
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)

def clean_line(s):
    s = re.sub(r'[•·•\U0001F535\U0001F534\U0001F7E5]', '', s)
    s = re.sub(r'\b(1st|2nd|3rd|Premium)\b', '', s, flags=re.IGNORECASE)
    return s.strip(' \n\r\t·•-,')

def extract_cards(driver):
    return driver.execute_script("""
        var results = [];
        var links = document.querySelectorAll('a[href*="/in/"]');
        var seen = new Set();
        for (var link of links) {
            var href = link.href.split('?')[0].replace(/\\/+$/, '');
            if (!href.match(/linkedin\\.com\\/in\\/[^/]+$/) || seen.has(href)) continue;
            var node = link;
            var cardText = '';
            for (var i = 0; i < 8; i++) {
                node = node.parentElement;
                if (!node) break;
                var t = (node.innerText || '').trim();
                if (t.length > 20) { cardText = t; break; }
            }
            seen.add(href);
            results.push({url: href, text: cardText.substring(0, 400)});
        }
        return results;
    """)

def parse_card(card):
    url  = card['url']
    text = card['text']
    lines = [clean_line(l) for l in text.splitlines()]
    lines = [l for l in lines if l and len(l) > 1]

    if not lines:
        return None

    # Skip feed posts / non-profile cards
    skip_phrases = ['Feed post', 'commented on', 'reposted', 'Connect\n', 'Message\n', 'Follow\n']
    if any(p in text for p in skip_phrases) and len(lines) < 4:
        return None

    name_raw = lines[0]

    # Skip "mutual connections" noise lines
    if 'mutual connection' in name_raw.lower() or name_raw.count(',') >= 2:
        return None

    parts = name_raw.split(' ', 1)
    first = parts[0]
    last  = parts[1] if len(parts) > 1 else ''

    # First word must look like a real name (letters only, not a number)
    if not first or not re.match(r'^[A-Za-z֐-׿؀-ۿ]+', first):
        return None

    job_title = ''
    company   = ''

    for line in lines[1:4]:
        if ' at ' in line:
            job_title, company = line.split(' at ', 1)
            break
        elif line and not re.match(r'^[\d\s]+h?$', line):
            if not job_title:
                job_title = line

    return {
        'firstName': first.strip(),
        'lastName': last.strip(),
        'linkedinProfileUrl': url,
        'jobTitle': job_title.strip(),
        'company': company.strip(),
    }

def scrape_page(driver, existing_urls):
    time.sleep(random.uniform(3, 5))
    new_rows = []
    cards = extract_cards(driver)

    for card in cards:
        s = slug(card['url'])
        if card['url'] in existing_urls or (s and s in existing_urls):
            continue
        parsed = parse_card(card)
        if not parsed or not parsed['firstName']:
            continue
        if not is_tech(parsed):
            continue
        existing_urls.add(card['url'])
        new_rows.append(parsed)

    return new_rows

def scrape_keyword(driver, keyword, existing_urls):
    all_new = []
    encoded = quote(keyword)
    print(f"\n>>> Keyword: {keyword}")

    for page in range(1, MAX_PAGES_PER_KEYWORD + 1):
        url = (
            f"https://www.linkedin.com/search/results/people/"
            f"?keywords={encoded}"
            f"&geoUrn=%5B%22{ISRAEL_GEO}%22%5D"
            f"&heroType=HIRING"
            f"&origin=FACETED_SEARCH"
            f"&page={page}"
        )
        driver.get(url)
        rows = scrape_page(driver, existing_urls)

        if not rows:
            print(f"  Page {page}: no new results - stopping")
            break

        all_new.extend(rows)
        print(f"  Page {page}: {len(rows)} new profiles")
        time.sleep(random.uniform(2, 3))

    return all_new

def main():
    print("Connecting to Chrome...")
    driver = connect()

    done_keywords = load_done()
    existing_urls = load_existing_urls()
    print(f"Existing profiles: {len(existing_urls)}")

    total_new = 0
    for keyword in KEYWORDS:
        if total_new >= MAX_TOTAL:
            print(f"\nReached {MAX_TOTAL} profiles limit — stopping.")
            break

        if keyword in done_keywords:
            print(f"Skipping '{keyword}' - already scraped")
            continue

        new_rows = scrape_keyword(driver, keyword, existing_urls)
        if new_rows:
            remaining = MAX_TOTAL - total_new
            new_rows = new_rows[:remaining]
            append_rows(new_rows)
            total_new += len(new_rows)
            print(f"  Saved {len(new_rows)} new profiles (total: {total_new}/{MAX_TOTAL})")

        save_done(keyword)
        time.sleep(random.uniform(5, 8))

    print(f"\nDone. Total new profiles: {total_new}")
    print(f"File: {OUTPUT_FILE}")
    import_to_xlsx()

def import_to_xlsx():
    if not os.path.exists(OUTPUT_FILE):
        return
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    li_col = headers.index('LinkedIn URL') + 1

    existing_slugs = set()
    for i in range(2, ws.max_row + 1):
        s = slug(ws.cell(i, li_col).value)
        if s: existing_slugs.add(s)

    added = 0
    with open(OUTPUT_FILE, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            url = row.get('linkedinProfileUrl', '').strip()
            s = slug(url)
            if not s or s in existing_slugs:
                continue
            new_row = [''] * len(headers)
            new_row[headers.index('First Name')]  = row.get('firstName', '')
            new_row[headers.index('Last Name')]   = row.get('lastName', '')
            new_row[headers.index('Company')]     = row.get('company', '')
            if 'Job Title' in headers:
                new_row[headers.index('Job Title')] = row.get('jobTitle', '')
            new_row[headers.index('LinkedIn URL')] = url.split('?')[0].rstrip('/')
            new_row[headers.index('Status')]      = 'Pending'
            ws.append(new_row)
            existing_slugs.add(s)
            added += 1

    wb.save(XLSX)
    print(f"Imported {added} new recruiters to outreach_list.xlsx")

if __name__ == '__main__':
    main()
