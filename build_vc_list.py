"""
Builds vc_outreach_list.xlsx - Israeli VC funds outreach list.

Emails marked VERIFIED were read off the fund's own website.
Funds with no published email are left with a blank Email cell and
Status = "NeedContact" - send_vc_emails.py will never send to a blank
or unverified address, so nothing goes out to a guessed mailbox.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = r"C:\Users\Shaul\Documents\job-search\vc_outreach_list.xlsx"

HEADERS = [
    "Fund", "Website", "Focus", "Pitch Phrase", "Fit", "Contact Name", "Contact Title",
    "Email", "Email Source", "Contact URL", "Jobs Board", "Status", "Date Sent", "Notes",
]

# Short, natural phrase dropped into the letter:
#   "You back a lot of {pitch} companies, and that's the world I come from."
# Keep it lowercase and readable out loud. Two areas maximum.
PITCH = {
    "Team8": "cybersecurity and enterprise data",
    "TLV Partners": "cyber and deep tech",
    "Pitango VC": "AI and deep tech",
    "Hetz Ventures": "dev tools and AI infrastructure",
    "Viola Ventures": "enterprise software and cyber",
    "F2 Venture Capital": "AI and deep tech",
    "StageOne Ventures": "deep tech and enterprise infrastructure",
    "Glilot Capital Partners": "cyber and enterprise software",
    "JVP (Jerusalem Venture Partners)": "cyber and AI",
    "Hanaco Ventures": "enterprise software and fintech",
    "New Era Capital Partners": "deep tech",
    "Cardumen Capital": "early stage tech",
    "Pontifax": "healthcare technology",
    "Grove Ventures": "AI and cloud infrastructure",
    "Aleph": "SaaS and fintech",
    "Cyberstarts": "cybersecurity",
    "Vertex Ventures Israel": "enterprise software and cyber",
    "State of Mind Ventures": "deep tech and enterprise software",
    "10D": "early stage AI",
    "Entree Capital": "AI and enterprise software",
    "Hyperwise Ventures": "cyber and deep tech",
    "Magma Venture Partners": "software and communications",
    "Greenfield Partners": "enterprise software",
    "Qumra Capital": "growth stage software",
    "Vintage Investment Partners": "growth stage technology",
    "Amiti Ventures": "deep tech and enterprise software",
    "Jibe Ventures": "deep tech and enterprise software",
    "S Capital": "enterprise software and cyber",
    "Swish Ventures": "enterprise software and cloud",
    "Magenta Venture Partners": "deep tech and enterprise software",
    "AnD Ventures": "deep tech and enterprise software",
    "OTV (Olive Tree Ventures)": "digital health",
    "TPY Capital": "early stage tech",
    "Firstime Venture Capital": "early stage tech",
    "Moneta VC": "fintech and enterprise software",
    "iAngels": "early stage tech",
    "Israel Growth Partners (IGP)": "growth stage software",
    "Maniv Mobility": "mobility and smart cities",
    "OurCrowd": "early stage tech",
    "Elron Ventures": "deep tech and cyber",
    "Gemini Israel Ventures": "early stage tech",
    "Crescendo Venture Partners": "deep tech",
    "Aristagora VC": "early stage tech",
    "Symbol": "early stage tech",
    "FLORA Ventures": "early stage tech",
    "BuiltUp Ventures": "early stage tech",
    "ION Crossover Partners": "late stage technology",
    "Israel Secondary Fund (ISF)": "technology",
    "ITI Venture Capital Partners": "technology",
    "vgames": "gaming",
    "Elliptic Ventures": "gaming and deep tech",
    "Beyond22": "ecommerce",
    "HIVE 2040": "fintech and cyber",
    "AltaIR VC": "SaaS and cyber",
    "Welltech Ventures": "AI",
    "Alliance Ventures (Renault-Nissan-Mitsubishi)": "mobility and autonomous systems",
    "Capital Nature": "cleantech and energy",
    "aMoon": "health technology",
    "Clal Biotechnology Industries": "biotech",
    "Triventures": "digital health",
    "Peregrine Ventures": "health technology",
    "eHealth Ventures": "digital health",
    "Joy Ventures": "neurotech",
    "Millennium Food-Tech": "foodtech",
    "IN Venture (Sumitomo)": "enterprise technology",
}

# Fit: High = enterprise SaaS / cyber / AI / deep tech (Shaul's domain)
#      Med  = adjacent (fintech, mobility, gaming infra)
#      Low  = biotech / medical / energy (weak match, keep for completeness)
ROWS = [
    # ---- VERIFIED EMAILS -------------------------------------------------
    ["Team8", "team8.vc", "Cyber, Enterprise, Data, Fintech, AI", "High", "", "", "info@team8.vc", "VERIFIED", "https://team8.vc/contact", "", "Pending", "", "Company-builder model - strong talent function"],
    ["TLV Partners", "tlv.partners", "Cyber, Deep Tech, Biotech", "High", "", "", "tlv@tlv.partners", "VERIFIED", "https://tlv.partners", "", "Pending", "", ""],
    ["Pitango VC", "pitango.com", "AI, Deep Tech, HealthTech, Fintech", "High", "", "", "mail@pitango.com", "VERIFIED", "https://pitango.com/contact", "", "Pending", "", "$2.8B AUM, 250+ companies"],
    ["Hetz Ventures", "hetz.vc", "Dev Tools, AI, Cyber, Deep Tech", "High", "", "", "contact@hetz.vc", "VERIFIED", "https://hetz.vc", "", "Pending", "", "Seed/early - dev-tools heavy"],
    ["Viola Ventures", "viola.vc", "AI, Cyber, Enterprise, Insurtech", "High", "", "", "info@viola.vc", "VERIFIED", "https://viola-group.com", "", "Pending", "", "Viola Group - largest IL tech platform"],
    ["F2 Venture Capital", "f2vc.com", "AI, Deep Tech, Cyber, Fintech", "High", "", "", "contact@f2vc.com", "VERIFIED", "https://f2vc.com", "", "Pending", "", ""],
    ["StageOne Ventures", "stageonevc.com", "Deep Tech, Enterprise Infrastructure", "High", "", "", "info@stageonevc.com", "VERIFIED", "https://stageonevc.com", "", "Pending", "", ""],
    ["Glilot Capital Partners", "glilotcapital.com", "Cyber, Enterprise Software, AI", "High", "", "", "info@glilotcapital.com", "VERIFIED", "https://glilotcapital.com/contact", "", "Pending", "", "Contact form has a 'job seeker' category"],
    ["JVP (Jerusalem Venture Partners)", "jvpvc.com", "Cyber, Fintech, AI, Agritech", "High", "", "", "info@jvpvc.com", "VERIFIED", "https://jvpvc.com", "https://jobs.jvpvc.com/jobs", "Pending", "", "Has a live portfolio jobs board - check it too"],
    ["Hanaco Ventures", "hanacovc.com", "Enterprise, Fintech, Cyber", "High", "", "", "hanaco@hanaco.com", "VERIFIED", "https://hanacovc.com/contact", "", "Pending", "", ""],
    ["New Era Capital Partners", "neweracp.com", "Deep Tech, Growth Stage", "High", "", "", "office@neweracp.com", "VERIFIED", "https://neweracp.com", "", "Pending", "", ""],
    ["Cardumen Capital", "cardumencapital.com", "Early-stage Tech, Israel/Europe", "Med", "", "", "info@cardumencapital.com", "VERIFIED", "https://cardumencapital.com", "", "Pending", "", ""],
    ["Pontifax", "pontifax.com", "Bio-pharma, Healthcare", "Low", "", "", "info@pontifax.com", "VERIFIED", "https://pontifax.com", "", "Pending", "", "Weak fit - medical only"],

    # ---- NEED CONTACT (no public email; use contact form or LinkedIn) -----
    ["Grove Ventures", "grovevc.com", "AI, Deep Tech, Cloud Infra", "High", "", "", "", "NEEDS_LOOKUP", "https://grovevc.com/contact/", "", "NeedContact", "", "Email obfuscated on site - grab from contact page or LinkedIn"],
    ["Aleph", "aleph.vc", "Fintech, SaaS, Mobile, EdTech", "High", "", "", "", "NEEDS_LOOKUP", "https://aleph.vc/about", "", "NeedContact", "", "Has 'Ampliphy' value-gen platform - that team handles talent"],
    ["Cyberstarts", "cyberstarts.com", "Cyber (seed)", "High", "", "", "", "NEEDS_LOOKUP", "https://cyberstarts.com", "", "NeedContact", "", "Very strong cyber portfolio; contact via LinkedIn"],
    ["Vertex Ventures Israel", "vertexventures.co.il", "Enterprise, Cyber, AI, SaaS", "High", "", "", "", "NEEDS_LOOKUP", "https://vertexventures.co.il", "", "NeedContact", "", "196 investments, 51 exits"],
    ["State of Mind Ventures", "somv.com", "Deep Tech, Enterprise, Cyber", "High", "", "", "", "NEEDS_LOOKUP", "https://somv.com", "", "NeedContact", "", ""],
    ["10D", "10d.vc", "Early-stage Tech, AI", "High", "", "", "", "NEEDS_LOOKUP", "https://10d.vc", "", "NeedContact", "", ""],
    ["Entree Capital", "entreecap.com", "AI, Cyber, Fintech, Big Data, SaaS", "High", "", "", "", "NEEDS_LOOKUP", "https://entreecap.com", "", "NeedContact", "", "281 investments - very broad portfolio"],
    ["Hyperwise Ventures", "hyperwise.vc", "Cyber, Deep Tech (seed)", "High", "", "", "", "NEEDS_LOOKUP", "https://hyperwise.vc", "", "NeedContact", "", ""],
    ["Magma Venture Partners", "magmavc.com", "ICT, Software, New Media", "High", "", "", "", "NEEDS_LOOKUP", "https://magmavc.com", "", "NeedContact", "", ""],
    ["Greenfield Partners", "greenfield-growth.com", "Enterprise Software, Growth", "High", "", "", "", "NEEDS_LOOKUP", "https://greenfield-growth.com", "", "NeedContact", "", "Growth stage - companies already scaling QA"],
    ["Qumra Capital", "qumracapital.com", "Growth Stage Tech", "High", "", "", "", "NEEDS_LOOKUP", "https://qumracapital.com", "", "NeedContact", "", "Late stage - real QA org needs"],
    ["Vintage Investment Partners", "vintage-ip.com", "Fund-of-funds, Growth, Secondaries", "High", "", "", "", "NEEDS_LOOKUP", "https://vintage-ip.com", "", "NeedContact", "", "Huge indirect network - good referral source"],
    ["Amiti Ventures", "amiti.vc", "Deep Tech, Enterprise", "High", "", "", "", "NEEDS_LOOKUP", "https://amiti.vc", "", "NeedContact", "", ""],
    ["Jibe Ventures", "jibevc.com", "Deep Tech, Enterprise", "High", "", "", "", "NEEDS_LOOKUP", "https://jibevc.com", "", "NeedContact", "", ""],
    ["S Capital", "scapitalvc.com", "Enterprise, Cyber, AI", "High", "", "", "", "NEEDS_LOOKUP", "https://scapitalvc.com", "", "NeedContact", "", ""],
    ["Swish Ventures", "swish.vc", "Enterprise Software, Cloud", "High", "", "", "", "NEEDS_LOOKUP", "https://swish.vc", "", "NeedContact", "", ""],
    ["Magenta Venture Partners", "magenta.vc", "Deep Tech, Enterprise", "High", "", "", "", "NEEDS_LOOKUP", "https://magenta.vc", "", "NeedContact", "", ""],
    ["AnD Ventures", "and-ventures.com", "Deep Tech, Enterprise", "High", "", "", "", "NEEDS_LOOKUP", "https://and-ventures.com", "", "NeedContact", "", ""],
    ["OTV (Olive Tree Ventures)", "otv.vc", "Digital Health, Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://otv.vc", "", "NeedContact", "", ""],
    ["TPY Capital", "tpycapital.com", "Early-stage Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://tpycapital.com", "", "NeedContact", "", ""],
    ["Firstime Venture Capital", "firstime.vc", "Multi-sector Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://firstime.vc", "", "NeedContact", "", ""],
    ["Moneta VC", "monetavc.com", "Fintech, Enterprise", "Med", "", "", "", "NEEDS_LOOKUP", "https://monetavc.com", "", "NeedContact", "", ""],
    ["iAngels", "iangels.com", "Multi-sector Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://iangels.com", "", "NeedContact", "", ""],
    ["Israel Growth Partners (IGP)", "igpcapital.com", "Growth Stage Software", "Med", "", "", "", "NEEDS_LOOKUP", "https://igpcapital.com", "", "NeedContact", "", ""],
    ["Maniv Mobility", "maniv.com", "Mobility, Smart Cars, Smart Cities", "Med", "", "", "", "NEEDS_LOOKUP", "https://maniv.com", "", "NeedContact", "", "Safety-critical software - QA rigor matters"],
    ["OurCrowd", "ourcrowd.com", "Multi-sector (equity crowdfunding)", "Med", "", "", "", "NEEDS_LOOKUP", "https://ourcrowd.com", "", "NeedContact", "", "225k investors, very large portfolio"],
    ["Elron Ventures", "elron.com", "Deep Tech, Medical, Cyber", "Med", "", "", "", "NEEDS_LOOKUP", "https://elron.com", "", "NeedContact", "", ""],
    ["Gemini Israel Ventures", "gemini.co.il", "Multi-sector Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://gemini.co.il", "", "NeedContact", "", ""],
    ["Crescendo Venture Partners", "cr-vp.com", "Deep Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://cr-vp.com", "", "NeedContact", "", ""],
    ["Aristagora VC", "aristagoravc.com", "Early-stage Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://aristagoravc.com", "", "NeedContact", "", ""],
    ["Symbol", "symbol.vc", "Early-stage Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://symbol.vc", "", "NeedContact", "", ""],
    ["FLORA Ventures", "floravc.com", "Early-stage Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://floravc.com", "", "NeedContact", "", ""],
    ["BuiltUp Ventures", "builtupventures.com", "Early-stage Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://builtupventures.com", "", "NeedContact", "", ""],
    ["ION Crossover Partners", "ion-am.com", "Crossover / Late Stage", "Med", "", "", "", "NEEDS_LOOKUP", "https://ion-am.com", "", "NeedContact", "", ""],
    ["Israel Secondary Fund (ISF)", "israelsecondary.com", "Secondaries", "Low", "", "", "", "NEEDS_LOOKUP", "https://israelsecondary.com", "", "NeedContact", "", ""],
    ["ITI Venture Capital Partners", "itivcp.com", "Multi-sector", "Low", "", "", "", "NEEDS_LOOKUP", "https://itivcp.com", "", "NeedContact", "", ""],
    ["vgames", "vgames.vc", "Gaming", "Low", "", "", "", "NEEDS_LOOKUP", "https://vgames.vc", "", "NeedContact", "", ""],
    ["Elliptic Ventures", "", "Gaming, Mobile & Social, Deep Tech", "Low", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["Beyond22", "", "eCommerce", "Low", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["HIVE 2040", "", "Fintech, IoT, Smart Cities, Cyber, AI", "Med", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["AltaIR VC", "", "AdTech, SaaS, Cyber, AI, EdTech, Fintech", "High", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["Welltech Ventures", "", "AI", "Med", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["Alliance Ventures (Renault-Nissan-Mitsubishi)", "", "Mobility, IoT, Cyber, AI, Autonomous", "Med", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["Capital Nature", "", "CleanTech, Energy, Industrial", "Low", "", "", "", "NEEDS_LOOKUP", "", "", "NeedContact", "", "From lastartup list"],
    ["aMoon", "amoonfund.com", "Medical, HealthTech", "Low", "", "", "", "NEEDS_LOOKUP", "https://amoonfund.com", "", "NeedContact", "", "Weak fit"],
    ["Clal Biotechnology Industries", "cbi-israel.com", "Biotech, Medical", "Low", "", "", "", "NEEDS_LOOKUP", "https://cbi-israel.com", "", "NeedContact", "", "Weak fit"],
    ["Triventures", "triventures.vc", "Digital Health, MedTech", "Low", "", "", "", "NEEDS_LOOKUP", "https://triventures.vc", "", "NeedContact", "", "Weak fit"],
    ["Peregrine Ventures", "peregrinevc.com", "Medical, HealthTech", "Low", "", "", "", "NEEDS_LOOKUP", "https://peregrinevc.com", "", "NeedContact", "", "Weak fit"],
    ["eHealth Ventures", "ehealthventures.com", "Digital Health", "Low", "", "", "", "NEEDS_LOOKUP", "https://ehealthventures.com", "", "NeedContact", "", "Weak fit"],
    ["Joy Ventures", "joyventures.com", "Neurotech, Wellbeing", "Low", "", "", "", "NEEDS_LOOKUP", "https://joyventures.com", "", "NeedContact", "", "Weak fit"],
    ["Millennium Food-Tech", "millennium-ft.com", "FoodTech", "Low", "", "", "", "NEEDS_LOOKUP", "https://millennium-ft.com", "", "NeedContact", "", "Weak fit"],
    ["IN Venture (Sumitomo)", "in-venture.com", "Multi-sector Tech", "Med", "", "", "", "NEEDS_LOOKUP", "https://in-venture.com", "", "NeedContact", "", ""],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "VC Outreach"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    verified_fill = PatternFill("solid", fgColor="E2EFDA")   # green
    needs_fill = PatternFill("solid", fgColor="FFF2CC")      # amber

    # ROWS is authored without the Pitch Phrase column; splice it in after Focus
    # so the source list stays readable and the two can't drift apart.
    missing_pitch = [r[0] for r in ROWS if r[0] not in PITCH]
    if missing_pitch:
        raise SystemExit(f"No pitch phrase defined for: {missing_pitch}")

    for row in ROWS:
        full = row[:3] + [PITCH[row[0]]] + row[3:]
        ws.append(full)
        fill = verified_fill if full[8] == "VERIFIED" else needs_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    widths = [34, 26, 40, 34, 6, 16, 18, 30, 14, 38, 30, 12, 12, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)

    verified = sum(1 for r in ROWS if r[7] == "VERIFIED")
    high = sum(1 for r in ROWS if r[3] == "High")
    print(f"Wrote {OUT}")
    print(f"  Total funds:      {len(ROWS)}")
    print(f"  Verified emails:  {verified}  (Status=Pending, ready to send)")
    print(f"  Need lookup:      {len(ROWS) - verified}  (Status=NeedContact)")
    print(f"  High-fit funds:   {high}")


if __name__ == "__main__":
    main()
