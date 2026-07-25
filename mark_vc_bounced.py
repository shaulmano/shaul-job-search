"""
Marks bounced addresses in vc_outreach_list.xlsx.

Edits the live file in place so Sent statuses and dates survive.
Do NOT run build_vc_list.py to fix bounces - it rebuilds the sheet from
scratch and would wipe everything you have already sent.

Usage:
    python mark_vc_bounced.py info@viola.vc contact@f2vc.com hanaco@hanaco.com
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

VC_FILE = r"C:\Users\Shaul\Documents\job-search\vc_outreach_list.xlsx"


def main():
    bad = {a.strip().lower() for a in sys.argv[1:]}
    if not bad:
        raise SystemExit("Pass the bounced addresses as arguments.")

    wb = load_workbook(VC_FILE)
    ws = wb.active
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}

    found = set()
    for row in ws.iter_rows(min_row=2):
        email = str(row[headers["Email"]].value or "").strip().lower()
        if email not in bad:
            continue
        found.add(email)
        fund = row[headers["Fund"]].value
        row[headers["Status"]].value = "Bounced"
        row[headers["Email Source"]].value = "BOUNCED"
        note = str(row[headers["Notes"]].value or "").strip()
        addition = f"{email} bounced, address is dead, needs a new one"
        row[headers["Notes"]].value = f"{note} | {addition}" if note else addition
        print(f"  marked bounced: {fund} <{email}>")

    for missing in bad - found:
        print(f"  WARNING: {missing} not found in the sheet")

    wb.save(VC_FILE)
    print(f"\nSaved {VC_FILE}")
    print("These rows no longer have Status=Pending, so they are excluded from sending.")


if __name__ == "__main__":
    main()
