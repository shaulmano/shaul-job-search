import sys
sys.stdout.reconfigure(encoding='utf-8')

import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
from openpyxl import load_workbook
from config import (
    GMAIL_ADDRESS, GMAIL_APP_PASSWORD, SENDER_NAME,
    SENDER_PHONE, SENDER_LINKEDIN, PDF_ATTACHMENT, CV_PM, CV_QA, OUTREACH_FILE
)

PREVIEW_COUNT = 5
EMAIL_SUBJECT = "25 Years in QA & Project Manager - Worth a 15-Minute Call?"

EMAIL_BODY = """\
Hi {first_name},

I'm reaching out because I believe my background could be a strong fit for {company}.

I'm Shaul Mano — a Senior Program Manager and QA Leader with 25+ years of experience at companies like RSA and Symantec.

If there's anything relevant open at {company}, I'd love a quick 15-minute call.
I've attached two CVs - one focused on Program Management and one on QA Leadership - so you can share whichever fits best.

Best,
Shaul Mano
{phone}
{linkedin}
"""


def build_email(to_address, first_name, company):
    msg = MIMEMultipart()
    msg["From"] = f"{SENDER_NAME} <{GMAIL_ADDRESS}>"
    msg["To"] = to_address
    msg["Subject"] = EMAIL_SUBJECT

    body = EMAIL_BODY.format(
        first_name=first_name,
        company=company,
        phone=SENDER_PHONE,
        linkedin=SENDER_LINKEDIN,
    )
    msg.attach(MIMEText(body, "plain"))

    for attachment in [PDF_ATTACHMENT, CV_PM, CV_QA]:
        if os.path.exists(attachment):
            with open(attachment, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f'attachment; filename="{os.path.basename(attachment)}"',
            )
            msg.attach(part)
        else:
            print(f"  WARNING: file not found: {attachment}")

    return msg


def send_email(msg, to_address):
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, to_address, msg.as_string())


def main():
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    pending = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status = row[5].value
        if status == "Pending":
            pending.append({
                "row_num": i,
                "first": str(row[0].value or "").strip(),
                "last": str(row[1].value or "").strip(),
                "company": str(row[2].value or "").strip(),
                "email": str(row[3].value or "").strip(),
            })

    if not pending:
        print("No pending emails to send.")
        return

    print(f"Found {len(pending)} pending emails.")
    daily_limit = int(input("How many to send today? (default 50): ").strip() or "50")
    batch = pending[:daily_limit]
    print()

    # Show preview of first 10
    preview_count = min(PREVIEW_COUNT, len(batch))
    print("=" * 60)
    print(f"PREVIEW — first {preview_count} of {len(batch)} emails:")
    print("=" * 60)
    for idx, item in enumerate(batch[:preview_count], 1):
        print(f"\n[{idx}] {item['first']} {item['last']} <{item['email']}> @ {item['company']}")
        print("-" * 60)
        body_preview = EMAIL_BODY.format(
            first_name=item["first"],
            company=item["company"],
            phone=SENDER_PHONE,
            linkedin=SENDER_LINKEDIN,
        )
        print(body_preview)
        print(f"Attachments: {os.path.basename(PDF_ATTACHMENT)} | {os.path.basename(CV_PM)} | {os.path.basename(CV_QA)}")
        print("=" * 60)

    answer = input(f"\nSend all {len(batch)} emails? (yes / no): ").strip().lower()
    if answer != "yes":
        print("Cancelled.")
        return

    print(f"\nSending {len(batch)} emails...\n")
    sent_count = 0
    for item in batch:
        try:
            msg = build_email(item["email"], item["first"], item["company"])
            send_email(msg, item["email"])
            ws.cell(row=item["row_num"], column=6).value = "Sent"
            ws.cell(row=item["row_num"], column=7).value = str(date.today())
            wb.save(OUTREACH_FILE)
            sent_count += 1
            print(f"  [{sent_count}/{len(batch)}] Sent to {item['first']} {item['last']} <{item['email']}>")
        except Exception as e:
            print(f"  ERROR sending to {item['email']}: {e}")

    print(f"\nDone. Sent {sent_count} / {len(batch)} emails.")


if __name__ == "__main__":
    main()
