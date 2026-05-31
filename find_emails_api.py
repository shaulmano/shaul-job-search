import sys
sys.stdout.reconfigure(encoding='utf-8')

import json
import time
import re
import smtplib
import threading
import urllib.request
import requests
import dns.resolver
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from config import (
    HUNTER_API_KEY, SNOV_CLIENT_ID, SNOV_CLIENT_SECRET, APOLLO_API_KEY, OUTREACH_FILE
)

PHANTOM_JSON_URL = "https://phantombuster.s3.amazonaws.com/vext4jKOYRY/4IcQxOVDThFpVMKnrjLSpA/result.json"
WORKERS = 3
_hunter_lock = threading.Lock()
_hunter_exhausted = {"value": False}

_domain_cache = {}
_domain_lock  = threading.Lock()
_print_lock   = threading.Lock()
_save_lock    = threading.Lock()
_snov_token   = {"value": None}
_snov_lock    = threading.Lock()

EMAIL_PATTERNS = [
    lambda f, l: f"{f}.{l}",
    lambda f, l: f"{f}",
    lambda f, l: f"{f[0]}{l}",
    lambda f, l: f"{f}{l[0]}",
    lambda f, l: f"{f}_{l}",
    lambda f, l: f"{f}{l}",
]

BAD_DOMAINS = {
    "linkedin.com", "facebook.com", "glassdoor.com", "indeed.com",
    "wikipedia.org", "zoominfo.com", "crunchbase.com", "bloomberg.com",
    "ynetnews.com", "bbc.co.uk", "haaretz.com", "calcalist.co.il",
    "youtube.com", "twitter.com", "instagram.com", "tiktok.com",
    "rocketreach.co", "rocketreach.io", "contactout.com", "finalscout.com",
    "leadiq.com", "apollo.io", "lusha.com", "clearbit.com", "hunter.io",
    "nytimes.com", "timesofisrael.com", "haaretz.com", "jpost.com",
    "medium.com", "substack.com", "reddit.com", "quora.com",
    "investing.com", "businesswire.com", "prnewswire.com",
    "theorg.com", "growjo.com", "beststartup.asia", "kycisrael.com",
    "slideshare.net", "docplayer.net", "scribd.com",
    "linktr.ee", "linkin.bio", "affplus.com",
    "volza.com", "dnb.com", "vault.com", "datanyze.com",
    "israelforever.org", "israelgives.org", "israelitimes.com",
    "calcalistech.com", "techrepublic.com", "talkbusiness.net",
    "drushim.co.il", "alljobs.co.il", "jobmaster.co.il",
    "github.com", "gitlab.com", "stackoverflow.com",
    "hbomax.com", "palestinechronicle.com", "cufi.org",
    "ritholtz.com", "gigaom.com", "ramp.directory",
    "mastersportal.com", "webatla.com", "italent.co.il",
    "intch.org", "ebenezer.org.il", "timetoast.com",
    "noatishby.com", "nortech-platform.com", "affplus.com",
}


def log(msg):
    with _print_lock:
        print(msg, flush=True)


# ── Domain lookup ─────────────────────────────────────────────────────────────

def get_mx(domain):
    try:
        records = dns.resolver.resolve(domain, 'MX', lifetime=5)
        return sorted(records, key=lambda r: r.preference)[0].exchange.to_text().rstrip('.')
    except Exception:
        return None


def _company_keywords(company):
    """Return set of meaningful words from company name for relevance check."""
    stop = {"ltd", "ltda", "inc", "llc", "gmbh", "co", "group", "the",
            "of", "and", "for", "by", "il", "israel", "technologies",
            "solutions", "systems", "services", "global", "digital"}
    words = re.findall(r'[a-z0-9]+', company.lower())
    return {w for w in words if len(w) >= 3 and w not in stop}


def _domain_relevant(domain, company):
    """Return True if domain plausibly belongs to the company."""
    domain_clean = re.sub(r'\.(com|io|ai|co|il|net|org|co\.il)$', '', domain.lower())
    domain_clean = re.sub(r'[^a-z0-9]', '', domain_clean)
    keywords = _company_keywords(company)
    # Accept if any keyword appears in the domain
    for kw in keywords:
        if kw in domain_clean:
            return True
    return False


def find_domain(company, slug=""):
    key = company.lower()
    with _domain_lock:
        if key in _domain_cache:
            return _domain_cache[key]

    result = ""

    # Try company slug guesses first (most reliable)
    if slug:
        clean = re.sub(r'[^a-z0-9]', '', slug.lower())
        for tld in [".com", ".io", ".ai", ".co", ".co.il"]:
            guess = clean + tld
            if get_mx(guess):
                result = guess
                break

    # Try company name directly as domain
    if not result:
        clean = re.sub(r'[^a-z0-9]', '', company.lower().split()[0])
        if len(clean) >= 3:
            for tld in [".com", ".co.il", ".io", ".ai"]:
                guess = clean + tld
                if get_mx(guess):
                    result = guess
                    break

    # DuckDuckGo fallback — require domain to be relevant
    if not result:
        try:
            from ddgs import DDGS
            for r in DDGS().text(f'"{company}" site:*.com OR site:*.io OR site:*.co.il', max_results=6):
                url = r.get("href", "")
                m = re.search(r'https?://(?:www\.)?([^/]+)', url)
                if not m:
                    continue
                domain = m.group(1).lower().split('/')[0]
                base = domain.split(':')[0]
                if base in BAD_DOMAINS:
                    continue
                if any(bad in base for bad in ["youtube", "linkedin", "facebook", "twitter",
                                                "rocketreach", "contactout", "nytimes",
                                                "timesofisrael", "jpost", "medium", "reddit"]):
                    continue
                if not _domain_relevant(base, company):
                    continue
                if get_mx(base):
                    result = base
                    break
            time.sleep(0.4)
        except Exception:
            time.sleep(0.5)

    with _domain_lock:
        _domain_cache[key] = result
    return result


# ── Hunter.io ─────────────────────────────────────────────────────────────────

def hunter_find(first, last, domain):
    if not HUNTER_API_KEY or _hunter_exhausted["value"]:
        return None
    with _hunter_lock:
        if _hunter_exhausted["value"]:
            return None
        try:
            resp = requests.get("https://api.hunter.io/v2/email-finder", params={
                "domain": domain, "first_name": first, "last_name": last,
                "api_key": HUNTER_API_KEY
            }, timeout=10)
            if resp.status_code == 200:
                data = resp.json().get("data", {})
                email = data.get("email")
                score = data.get("score", 0)
                if email and score >= 50:
                    return email
            elif resp.status_code in (429, 402):
                log("  [Hunter] quota exhausted — switching to SMTP only")
                _hunter_exhausted["value"] = True
        except Exception:
            pass
    return None


# ── Snov.io ───────────────────────────────────────────────────────────────────

def _get_snov_token():
    with _snov_lock:
        if _snov_token["value"]:
            return _snov_token["value"]
        if not SNOV_CLIENT_ID:
            return None
        try:
            resp = requests.post("https://api.snov.io/v1/oauth/access_token", data={
                "grant_type": "client_credentials",
                "client_id": SNOV_CLIENT_ID,
                "client_secret": SNOV_CLIENT_SECRET
            }, timeout=10)
            token = resp.json().get("access_token")
            _snov_token["value"] = token
            return token
        except Exception:
            return None


def snov_find(first, last, domain):
    token = _get_snov_token()
    if not token:
        return None
    try:
        requests.post("https://api.snov.io/v1/add-names-to-find-emails", data={
            "access_token": token, "firstName": first,
            "lastName": last, "domain": domain
        }, timeout=10)
        time.sleep(2)
        resp = requests.post("https://api.snov.io/v1/get-emails-from-names", data={
            "access_token": token, "firstName": first,
            "lastName": last, "domain": domain
        }, timeout=10)
        emails = resp.json().get("emails", [])
        if emails:
            best = max(emails, key=lambda e: e.get("confidence", 0))
            if best.get("confidence", 0) >= 50:
                return best.get("email")
    except Exception:
        pass
    return None


# ── Apollo.io ─────────────────────────────────────────────────────────────────

def apollo_find(first, last, company, linkedin_url):
    if not APOLLO_API_KEY:
        return None
    try:
        resp = requests.post("https://api.apollo.io/v1/people/match",
            headers={"x-api-key": APOLLO_API_KEY, "Content-Type": "application/json"},
            json={
                "first_name": first, "last_name": last,
                "organization_name": company,
                "linkedin_url": linkedin_url or None,
                "reveal_personal_emails": False
            }, timeout=10)
        if resp.status_code == 200:
            return resp.json().get("person", {}).get("email")
    except Exception:
        pass
    return None


# ── SMTP fallback ─────────────────────────────────────────────────────────────

def is_accept_all(mx):
    if not mx:
        return True
    return any(x in mx.lower() for x in ["google", "outlook", "microsoft", "protection.outlook"])


def smtp_verify(email, mx):
    try:
        with smtplib.SMTP(mx, 25, timeout=8) as s:
            s.ehlo("mail.test.com")
            s.mail("verify@test.com")
            code, _ = s.rcpt(email)
            return code == 250
    except Exception:
        return None


def smtp_fallback(first, last, domain):
    mx = get_mx(domain)
    candidates = [f"{p(first, last)}@{domain}" for p in EMAIL_PATTERNS]
    if is_accept_all(mx) or not mx:
        return candidates[0], False
    for email in candidates:
        if smtp_verify(email, mx) is True:
            return email, True
    return "", False


# ── Main processor ────────────────────────────────────────────────────────────

def normalize(name):
    return re.sub(r'[^a-z]', '', name.lower().strip())


def process_profile(idx, p, existing):
    first_raw = (p.get("firstName") or "").strip()
    last_raw  = (p.get("lastName")  or "").strip()
    company   = (p.get("company") or p.get("companyName") or "").strip()
    linkedin  = (p.get("profileUrl") or p.get("linkedinProfileUrl") or "").strip()
    slug      = (p.get("companySlug") or "").strip()

    first = normalize(first_raw)
    last  = normalize(last_raw)

    if not first or not last or not company:
        return None, "skip"
    if company.lower() in existing:
        return None, "existing"

    domain = find_domain(company, slug)
    if not domain:
        log(f"  [{idx}] {first_raw} {last_raw} @ {company} → no domain")
        return None, "no_domain"

    source = ""

    # 1. Hunter.io
    email = hunter_find(first, last, domain)
    if email:
        source = "Hunter"

    # 2. Snov.io
    if not email:
        email = snov_find(first, last, domain)
        if email:
            source = "Snov"

    # 3. Apollo.io
    if not email:
        email = apollo_find(first_raw, last_raw, company, linkedin)
        if email:
            source = "Apollo"

    # 4. SMTP fallback
    if not email:
        email, verified = smtp_fallback(first, last, domain)
        if email:
            source = "SMTP✓" if verified else "SMTP~"

    if not email:
        log(f"  [{idx}] {first_raw} {last_raw} @ {company} → no email")
        return None, "no_email"

    # Validate format
    if not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
        log(f"  [{idx}] {first_raw} {last_raw} @ {company} → invalid email: {email}")
        return None, "no_email"

    log(f"  [{idx}] {first_raw} {last_raw} @ {company} → {email} [{source}]")
    return {"first": first_raw, "last": last_raw, "company": company,
            "email": email, "linkedin": linkedin}, "ok"


# ── Save / Load ───────────────────────────────────────────────────────────────

def save_to_outreach(results):
    with _save_lock:
        try:
            wb = load_workbook(OUTREACH_FILE)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["First Name", "Last Name", "Company", "Email", "LinkedIn URL", "Status", "Date Sent"])
        for r in results:
            ws.append([r["first"], r["last"], r["company"], r["email"], r["linkedin"], "Pending", ""])
        wb.save(OUTREACH_FILE)


def load_existing():
    try:
        wb = load_workbook(OUTREACH_FILE)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3]:
                existing.add(str(row[3]).strip().lower())
            if row[2]:
                existing.add(str(row[2]).strip().lower())
        return existing
    except Exception:
        return set()


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    active = []
    if HUNTER_API_KEY:   active.append("Hunter.io")
    if SNOV_CLIENT_ID:   active.append("Snov.io")
    if APOLLO_API_KEY:   active.append("Apollo.io")
    active.append("SMTP fallback")
    log(f"Sources: {' → '.join(active)}\n")

    log("Downloading PhantomBuster profiles...")
    with urllib.request.urlopen(PHANTOM_JSON_URL, timeout=30) as r:
        profiles = json.loads(r.read())
    log(f"Total profiles: {len(profiles)}")

    existing = load_existing()
    existing_lock = threading.Lock()
    log(f"Already in outreach list: {len(existing)} entries\n")

    stats = {"added": 0, "existing": 0, "no_domain": 0, "no_email": 0, "skip": 0}
    batch = []
    batch_lock = threading.Lock()

    def run(args):
        idx, p = args
        result, status = process_profile(idx, p, existing)
        if result:
            with existing_lock:
                existing.add(result["company"].lower())
                existing.add(result["email"].lower())
            with batch_lock:
                batch.append(result)
                if len(batch) >= 10:
                    to_save = batch[:]
                    batch.clear()
                    save_to_outreach(to_save)
                    log(f"  >>> Saved 10 rows")
        with _print_lock:
            stats[status if status in stats else "added"] = stats.get(status, 0)
            if status == "ok":
                stats["added"] += 1
        return status

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = {ex.submit(run, (i+1, p)): i for i, p in enumerate(profiles)}
        done = 0
        for f in as_completed(futures):
            f.result()
            done += 1
            if done % 100 == 0:
                log(f"\n--- {done}/{len(profiles)} processed | found: {stats['added']} ---\n")

    if batch:
        save_to_outreach(batch)

    log(f"\n=== Done ===")
    log(f"Added:      {stats['added']}")
    log(f"No domain:  {stats['no_domain']}")
    log(f"No email:   {stats['no_email']}")


if __name__ == "__main__":
    main()
