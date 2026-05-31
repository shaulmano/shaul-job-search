import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

OUTREACH_FILE = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"

bounced_emails = {
    "hadar@pentera.io",
    "julie@gutsy.com",
    "ronni@hunters.ai",
    "rachel@loom.com",
    "nadia@letsdeel.com",
    "andrs@leapsome.com",
    "jan@zencoder.ai",
    "itamar@beehero.io",
    "catherine@kustomer.com",
    "april@totango.com",
    "jamie@joinblink.com",
    "maayan@moonactive.com",
    "timothy@workiz.com",
    "dhyey@cast.ai",
    "tal@glassbox.com",
    "t.grinshtein@sunbit.com",
    "yona@mobileye.com",
    "rotem@vesttoo.com",
    "lital@intezer.com",
    "dan@tipalti.com",
    "andy@nextinsurance.com",
}

wb = load_workbook(OUTREACH_FILE)
ws = wb.active

updated = 0
for row in ws.iter_rows(min_row=2):
    email = str(row[3].value or "").strip().lower()
    if email in bounced_emails:
        row[5].value = "Bounced"
        updated += 1
        print(f"  Marked Bounced: {email}")

wb.save(OUTREACH_FILE)
print(f"\nDone: {updated} rows marked as Bounced")
