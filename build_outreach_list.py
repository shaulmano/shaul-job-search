import sys
sys.stdout.reconfigure(encoding='utf-8')

import time
import re
import json
import urllib.request
import urllib.parse
from ddgs import DDGS
from openpyxl import load_workbook, Workbook

INPUT_FILE  = "C:\\Users\\Shaul\\Documents\\job-search\\companies.xlsx"
OUTPUT_FILE = "C:\\Users\\Shaul\\Documents\\job-search\\outreach_list.xlsx"

SNOV_USER_ID = "0cf74a351b5f8b55708e0bb76745a7ef"
SNOV_SECRET  = "103c4cc1874a49cfd582f86132fe5991"

_snov_token = None
_snov_token_expiry = 0
_snov_credits_ok = True


def snov_get_token():
    global _snov_token, _snov_token_expiry
    import time as t
    if _snov_token and t.time() < _snov_token_expiry:
        return _snov_token

    url = "https://api.snov.io/v1/oauth/access_token"
    data = json.dumps({
        "grant_type": "client_credentials",
        "client_id": SNOV_USER_ID,
        "client_secret": SNOV_SECRET,
    }).encode()

    try:
        req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            result = json.loads(r.read())
        _snov_token = result.get("access_token", "")
        _snov_token_expiry = t.time() + result.get("expires_in", 3600) - 60
        return _snov_token
    except Exception as e:
        print(f"  Snov.io auth error: {e}")
        return ""


def snov_request(method, url, body=None, token=None):
    """Make a Snov.io API request with Bearer auth."""
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


def snov_email_finder(first, last, domain):
    """Find a specific person's verified email via Snov.io (async)."""
    global _snov_credits_ok
    if not _snov_credits_ok:
        return ""

    token = snov_get_token()
    if not token:
        return ""

    try:
        # Step 1: start task
        body = {"rows": [{"first_name": first, "last_name": last, "domain": domain}]}
        start = snov_request("POST", "https://api.snov.io/v2/emails-by-domain-by-name/start", body, token)
        task_hash = start.get("data", {}).get("task_hash", "")
        if not task_hash:
            print(f"  Snov.io: no task_hash in response")
            return ""

        # Step 2: poll for result (max 5 attempts)
        result_url = f"https://api.snov.io/v2/emails-by-domain-by-name/result?task_hash={task_hash}"
        for attempt in range(5):
            time.sleep(2)
            result = snov_request("GET", result_url, token=token)
            status = result.get("status", "")

            if status == "completed":
                data_list = result.get("data", [])
                if data_list:
                    emails_found = data_list[0].get("result", [])
                    for e in emails_found:
                        email = e.get("email", "")
                        smtp = e.get("smtp_status", "")
                        if email and smtp in ("valid", "accept_all", ""):
                            print(f"  Snov.io: {email} ({smtp})")
                            return email
                return ""

            if "credit" in str(result).lower() or "limit" in str(result).lower():
                print(f"  Snov.io: credits exhausted")
                _snov_credits_ok = False
                return ""

        print(f"  Snov.io: task timeout after 5 attempts")
        return ""

    except Exception as e:
        print(f"  Snov.io email finder error: {e}")
        return ""


def snov_linkedin_lookup(linkedin_url):
    """Get profile info (and possibly email) from a LinkedIn URL via Snov.io."""
    global _snov_credits_ok
    if not _snov_credits_ok:
        return None

    token = snov_get_token()
    if not token:
        return None

    try:
        body = {"urls": [linkedin_url]}
        start = snov_request("POST", "https://api.snov.io/v2/li-profiles-by-urls/start", body, token)
        task_hash = start.get("data", {}).get("task_hash", "")
        if not task_hash:
            return None

        result_url = f"https://api.snov.io/v2/li-profiles-by-urls/result?task_hash={task_hash}"
        for _ in range(5):
            time.sleep(2)
            result = snov_request("GET", result_url, token=token)
            if result.get("status") == "completed":
                profiles = result.get("data", [])
                if profiles:
                    p = profiles[0]
                    email = p.get("email", "")
                    first = p.get("firstName") or p.get("first_name", "")
                    last  = p.get("lastName")  or p.get("last_name", "")
                    if email:
                        print(f"  Snov.io LinkedIn: {first} {last} <{email}>")
                    elif first:
                        print(f"  Snov.io LinkedIn profile: {first} {last} (no email)")
                    return {"first": first, "last": last, "email": email}
                return None
            if "credit" in str(result).lower():
                _snov_credits_ok = False
                return None
        return None
    except Exception as e:
        print(f"  Snov.io LinkedIn error: {e}")
        return None


def find_recruiter(company_name):
    """Find recruiter LinkedIn URL + name via DuckDuckGo."""
    queries = [
        f'linkedin.com/in "talent acquisition" "{company_name}" Israel',
        f'linkedin.com/in "recruiter" "{company_name}" Israel',
        f'linkedin.com/in "hr manager" "{company_name}" Israel',
    ]
    for query in queries:
        try:
            for r in DDGS().text(query, max_results=5):
                url = r.get("href", "")
                title = r.get("title", "")
                if "linkedin.com/in/" not in url:
                    continue
                name = title.split(" - ")[0].split(" | ")[0].strip()
                parts = name.split()
                if len(parts) >= 2 and not any(x in name.lower() for x in ["linkedin", "recruiter", "manager", "hr"]):
                    return parts[0], " ".join(parts[1:]), url
            time.sleep(1)
        except Exception:
            time.sleep(2)
            continue
    return "", "", ""


def load_companies():
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        domain = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if name and domain:
            companies.append({"name": name, "domain": domain})
    return companies


def load_existing_companies():
    import os
    if not os.path.exists(OUTPUT_FILE):
        return set()
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = str(row[2]).strip().lower() if row[2] else ""
        if company:
            existing.add(company)
    return existing


def save_batch(results):
    import os
    if not results:
        return
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Outreach List"
        ws.append(["First Name", "Last Name", "Company", "Email", "LinkedIn URL", "Status", "Date Sent"])
        for col, width in zip("ABCDEFG", [20, 20, 30, 38, 55, 12, 18]):
            ws.column_dimensions[col].width = width

    for r in results:
        ws.append([r["first"], r["last"], r["company"], r["email"], r["linkedin"], "Pending", ""])

    wb.save(OUTPUT_FILE)
    print(f"  >>> Saved {len(results)} rows to outreach_list.xlsx")


def main():
    print("Loading companies list...")
    companies = load_companies()
    existing = load_existing_companies()
    new_companies = [c for c in companies if c["name"].lower() not in existing]
    print(f"Total: {len(companies)} | Already done: {len(existing)} | Remaining: {len(new_companies)}\n")

    batch = []
    total_added = 0

    for i, company in enumerate(new_companies):
        name = company["name"]
        domain = company["domain"]
        print(f"[{i+1}/{len(new_companies)}] {name} ({domain})")

        # Step 1: Find recruiter name via DuckDuckGo
        first, last, linkedin_url = find_recruiter(name)
        time.sleep(1)

        if not first:
            print(f"  - Skipped: no recruiter found")
            continue

        print(f"  DuckDuckGo: {first} {last} | {linkedin_url}")

        # Step 2: Snov.io LinkedIn lookup — may return email directly + confirms name
        email = ""
        if linkedin_url:
            profile = snov_linkedin_lookup(linkedin_url)
            if profile:
                if profile.get("email"):
                    email = profile["email"]
                if profile.get("first"):
                    first = profile["first"]
                    last  = profile["last"]

        # Step 3: If no email yet, try Snov.io email finder with confirmed name
        if not email:
            email = snov_email_finder(first, last, domain)
            time.sleep(1)

        if not email:
            print(f"  - Skipped: email not found")
            continue

        batch.append({
            "first": first, "last": last,
            "company": name, "email": email, "linkedin": linkedin_url,
        })

        total_added += 1

        # Save every 5 results so we don't lose progress
        if len(batch) >= 5:
            save_batch(batch)
            batch = []

        time.sleep(2)

    # Save remaining
    if batch:
        save_batch(batch)

    print(f"\nDone. Added {total_added} new emails total.")


if __name__ == "__main__":
    main()
