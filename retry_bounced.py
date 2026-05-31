import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
import re

OUTREACH_FILE = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"

EMAIL_PATTERNS = [
    lambda f, l: f"{f}.{l}",    # shaul.mano
    lambda f, l: f"{f}",         # shaul
    lambda f, l: f"{f[0]}{l}",  # smano
    lambda f, l: f"{f}{l[0]}",  # shauls
    lambda f, l: f"{f}_{l}",    # shaul_mano
    lambda f, l: f"{f}{l}",     # Shaul
]


def normalize(name):
    return name.lower().strip().replace(" ", "").replace("-", "")


def get_domain(email):
    return email.split("@")[1] if "@" in email else ""


def local_part(email):
    return email.split("@")[0] if "@" in email else ""


def next_pattern(first, last, current_email):
    """Return next pattern email after the one that bounced."""
    f = normalize(first)
    l = normalize(last)
    domain = get_domain(current_email)
    current_local = local_part(current_email).lower()

    candidates = [f"{p(f, l)}@{domain}" for p in EMAIL_PATTERNS]

    # Find index of current pattern
    try:
        idx = [c.split("@")[0] for c in candidates].index(current_local)
        # Return next unused pattern
        for next_email in candidates[idx + 1:]:
            if next_email.lower() != current_email.lower():
                return next_email
    except ValueError:
        # Current pattern not recognized — start from beginning
        for c in candidates:
            if c.lower() != current_email.lower():
                return c

    return ""


def main():
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    # Collect existing emails to avoid duplicates
    existing_emails = set()
    bounced_rows = []

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        email = str(row[3].value or "").strip().lower()
        status = str(row[5].value or "").strip()
        if email:
            existing_emails.add(email)
        if status == "Bounced":
            bounced_rows.append({
                "row": i,
                "first": str(row[0].value or "").strip(),
                "last":  str(row[1].value or "").strip(),
                "company": str(row[2].value or "").strip(),
                "email": email,
                "linkedin": str(row[4].value or "").strip(),
            })

    print(f"Found {len(bounced_rows)} bounced emails\n")
    added = 0

    for item in bounced_rows:
        new_email = next_pattern(item["first"], item["last"], item["email"])
        if not new_email:
            print(f"  {item['email']} — no more patterns to try")
            continue
        # Validate email format
        if not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', new_email):
            print(f"  {item['email']} → {new_email} invalid format, skipping")
            continue
        if new_email.lower() in existing_emails:
            print(f"  {item['email']} → {new_email} already in list, skipping")
            continue

        ws.append([
            item["first"], item["last"], item["company"],
            new_email, item["linkedin"], "Pending", ""
        ])
        existing_emails.add(new_email.lower())
        added += 1
        print(f"  {item['email']}")
        print(f"  → {new_email} (added as Pending)")

    wb.save(OUTREACH_FILE)
    print(f"\nAdded {added} new retry rows.")


if __name__ == "__main__":
    main()
