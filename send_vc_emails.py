"""
Sends the VC outreach emails from vc_outreach_list.xlsx.

SAFETY - read this before running:
  * Default mode is DRY RUN. It prints every email in full and sends NOTHING.
        python send_vc_emails.py
  * To actually send you must pass --send AND type SEND at the prompt AND
    approve each individual email with y/n.
        python send_vc_emails.py --send
  * Only rows with Status == "Pending", a non-empty Email, and
    Email Source == "VERIFIED" are ever eligible. Guessed addresses
    (Status "NeedContact") are skipped - they bounce and burn sender reputation.
  * Default cap is 12 per run. Gmail starts flagging above ~15-20/day.

Columns are looked up BY HEADER NAME, not by index, so adding or moving a
column can't silently send to the wrong field.

Flags:
  --send            actually send (default is dry run)
  --he              use the Hebrew letter instead of English
  --limit N         cap this run at N emails (default 12)
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import smtplib
import argparse
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from openpyxl import load_workbook

from config import GMAIL_ADDRESS, GMAIL_APP_PASSWORD, SENDER_NAME

VC_FILE = r"C:\Users\Shaul\Documents\job-search\vc_outreach_list.xlsx"
CV_PATH = r"C:\Users\Shaul\Documents\CV\new\Shaul_Mano_CV_2026.pdf"
LETTER_EN = r"C:\Users\Shaul\Documents\job-search\vc_cover_letter_EN.txt"
LETTER_HE = r"C:\Users\Shaul\Documents\job-search\vc_cover_letter_HE.txt"


def load_letter(path):
    """Letter file starts with 'SUBJECT: ...' then a blank line, then the body."""
    with open(path, encoding="utf-8") as f:
        text = f.read()
    if not text.startswith("SUBJECT:"):
        raise SystemExit(f"{path} must start with a 'SUBJECT:' line")
    subject_line, _, body = text.partition("\n")
    return subject_line[len("SUBJECT:"):].strip(), body.strip()


def load_rows(ws):
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}
    required = ["Fund", "Pitch Phrase", "Email", "Email Source", "Status", "Contact Name"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"vc_outreach_list.xlsx is missing columns: {missing}")

    def col(row, name):
        return str(row[headers[name]].value or "").strip()

    eligible, skipped = [], 0
    for rownum, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if col(row, "Status") != "Pending":
            continue
        if not col(row, "Email") or col(row, "Email Source") != "VERIFIED":
            skipped += 1
            continue
        eligible.append({
            "row": rownum,
            "fund": col(row, "Fund"),
            "focus": col(row, "Pitch Phrase"),
            "email": col(row, "Email"),
            "name": col(row, "Contact Name"),
        })
    return eligible, skipped, headers


def render(body, item):
    name_part = f" {item['name'].split()[0]}" if item["name"] else " there"
    return body.format(name_part=name_part, fund=item["fund"], focus=item["focus"])


def build_msg(subject, body, item):
    msg = MIMEMultipart()
    msg["From"] = f"{SENDER_NAME} <{GMAIL_ADDRESS}>"
    msg["To"] = item["email"]
    msg["Subject"] = subject
    msg.attach(MIMEText(render(body, item), "plain", "utf-8"))

    if not os.path.exists(CV_PATH):
        raise SystemExit(f"CV not found: {CV_PATH}")
    with open(CV_PATH, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{os.path.basename(CV_PATH)}"')
    msg.attach(part)
    return msg


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--send", action="store_true", help="actually send (default: dry run)")
    ap.add_argument("--he", action="store_true", help="use the Hebrew letter")
    ap.add_argument("--limit", type=int, default=12)
    args = ap.parse_args()

    subject, body = load_letter(LETTER_HE if args.he else LETTER_EN)
    wb = load_workbook(VC_FILE)
    ws = wb.active
    eligible, skipped, headers = load_rows(ws)

    if not eligible:
        print("Nothing eligible to send. (Need Status=Pending + VERIFIED email.)")
        return

    batch = eligible[:args.limit]
    mode = "SEND" if args.send else "DRY RUN"
    print("=" * 70)
    print(f"MODE: {mode}   |   subject: {subject}")
    print(f"Eligible: {len(eligible)}   this run: {len(batch)}   "
          f"skipped (no verified email): {skipped}")
    print(f"Attachment: {os.path.basename(CV_PATH)}")
    print("=" * 70)

    if not args.send:
        for i, item in enumerate(batch, 1):
            print(f"\n[{i}] To: {item['email']}   ({item['fund']})")
            print("-" * 70)
            print(render(body, item))
            print("=" * 70)
        print("\nDRY RUN - nothing was sent. Re-run with --send when you're ready.")
        return

    if input('\nType SEND to confirm you want these going out for real: ').strip() != "SEND":
        print("Cancelled. Nothing sent.")
        return

    status_col = headers["Status"] + 1
    date_col = headers["Date Sent"] + 1
    sent = 0
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        for i, item in enumerate(batch, 1):
            print(f"\n[{i}/{len(batch)}] {item['fund']} <{item['email']}>")
            print("-" * 70)
            print(render(body, item))
            print("-" * 70)
            ans = input("Send this one? (y = send / n = skip / q = stop): ").strip().lower()
            if ans == "q":
                print("Stopped.")
                break
            if ans != "y":
                print("  skipped")
                continue
            try:
                msg = build_msg(subject, body, item)
                server.sendmail(GMAIL_ADDRESS, item["email"], msg.as_string())
                ws.cell(row=item["row"], column=status_col).value = "Sent"
                ws.cell(row=item["row"], column=date_col).value = str(date.today())
                wb.save(VC_FILE)
                sent += 1
                print(f"  SENT ({sent} so far)")
            except Exception as e:
                print(f"  ERROR: {e}")

    print(f"\nDone. Sent {sent} of {len(batch)}.")


if __name__ == "__main__":
    main()
