import sys
sys.stdout.reconfigure(encoding='utf-8')

import json
import time
import re
import smtplib
import threading
import urllib.request
import dns.resolver
from concurrent.futures import ThreadPoolExecutor, as_completed
from ddgs import DDGS
from openpyxl import load_workbook, Workbook

PHANTOM_JSON_URL = "https://phantombuster.s3.amazonaws.com/vext4jKOYRY/4IcQxOVDThFpVMKnrjLSpA/result.json"
OUTREACH_FILE = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"

WORKERS = 10          # parallel threads
DDG_SEMAPHORE = threading.Semaphore(3)  # max 3 DuckDuckGo searches at once

_domain_cache = {}
_domain_lock  = threading.Lock()
_print_lock   = threading.Lock()
_save_lock    = threading.Lock()
_counter      = {"done": 0, "added": 0, "total": 0}

EMAIL_PATTERNS = [
    lambda f, l: f"{f}.{l}",
    lambda f, l: f"{f}",
    lambda f, l: f"{f[0]}{l}",
    lambda f, l: f"{f}{l[0]}",
    lambda f, l: f"{f}_{l}",
    lambda f, l: f"{f}{l}",
]


def log(msg):
    with _print_lock:
        print(msg, flush=True)


def get_mx_server(domain):
    try:
        records = dns.resolver.resolve(domain, 'MX', lifetime=5)
        return sorted(records, key=lambda r: r.preference)[0].exchange.to_text().rstrip('.')
    except Exception:
        return None


def is_accept_all(mx):
    if not mx:
        return True
    return any(x in mx.lower() for x in ["google", "outlook", "microsoft", "protection.outlook"])


def smtp_verify(email, mx):
    try:
        with smtplib.SMTP(mx, 25, timeout=8) as s:
            s.ehlo("mail.test.com")
            s.mail("verify@test.com")
            code, _ = s.rcpt(email)
            if code == 250:
                return True
            if code in (550, 551, 553):
                return False
            return None
    except Exception:
        return None


def find_domain(company_name, slug=""):
    key = company_name.lower()
    with _domain_lock:
        if key in _domain_cache:
            return _domain_cache[key]

    # Try slug guesses first (no network rate limit needed)
    if slug:
        clean = slug.replace("-", "")
        for tld in [".com", ".io", ".ai", ".co"]:
            guess = clean + tld
            mx = get_mx_server(guess)
            if mx:
                with _domain_lock:
                    _domain_cache[key] = guess
                return guess

    # DuckDuckGo (rate-limited semaphore)
    with DDG_SEMAPHORE:
        try:
            query = f'"{company_name}" official site Israel'
            for r in DDGS().text(query, max_results=5):
                url = r.get("href", "")
                m = re.search(r'https?://(?:www\.)?([^/]+)', url)
                if not m:
                    continue
                domain = m.group(1).lower()
                if any(x in domain for x in ["linkedin", "facebook", "glassdoor", "indeed", "wikipedia", "jobsite"]):
                    continue
                mx = get_mx_server(domain)
                if mx:
                    with _domain_lock:
                        _domain_cache[key] = domain
                    return domain
            time.sleep(0.5)
        except Exception:
            time.sleep(1)

    with _domain_lock:
        _domain_cache[key] = ""
    return ""


def try_find_email(first, last, domain):
    first = first.lower().replace(" ", "").replace("-", "")
    last  = last.lower().replace(" ", "").replace("-", "")
    if not first or not last or not domain:
        return "", False

    mx = get_mx_server(domain)
    candidates = [f"{p(first, last)}@{domain}" for p in EMAIL_PATTERNS]

    if is_accept_all(mx) or not mx:
        return candidates[0], False

    for email in candidates:
        result = smtp_verify(email, mx)
        if result is True:
            return email, True
        elif result is False:
            pass
    return "", False


def process_profile(idx, p, existing):
    first   = (p.get("firstName") or "").strip()
    last    = (p.get("lastName")  or "").strip()
    company = (p.get("company") or p.get("companyName") or "").strip()
    linkedin_url = (p.get("profileUrl") or p.get("linkedinProfileUrl") or "").strip()
    slug    = (p.get("companySlug") or "").strip()

    if not first or not last or not company:
        return None, "skip"

    if company.lower() in existing:
        return None, "existing"

    domain = find_domain(company, slug)
    if not domain:
        log(f"  [{idx}] {first} {last} @ {company} → no domain")
        return None, "no_domain"

    email, verified = try_find_email(first, last, domain)
    if not email:
        log(f"  [{idx}] {first} {last} @ {company} ({domain}) → no email")
        return None, "no_email"

    tag = "✓" if verified else "~"
    log(f"  [{idx}] {first} {last} @ {company} → {email} {tag}")
    return {"first": first, "last": last, "company": company,
            "email": email, "linkedin": linkedin_url}, "ok"


def save_to_outreach(results):
    with _save_lock:
        try:
            wb = load_workbook(OUTREACH_FILE)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["First Name", "Last Name", "Company", "Email", "LinkedIn URL", "Status", "Date Sent"])
        for r in results:
            ws.append([r["first"], r["last"], r["company"], r["email"], r["linkedin"], "Pending", ""])
        wb.save(OUTREACH_FILE)


def load_existing():
    try:
        wb = load_workbook(OUTREACH_FILE)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3]:
                existing.add(str(row[3]).strip().lower())
            if row[2]:
                existing.add(str(row[2]).strip().lower())
        return existing
    except Exception:
        return set()


def main():
    log("Downloading PhantomBuster profiles...")
    with urllib.request.urlopen(PHANTOM_JSON_URL, timeout=30) as r:
        profiles = json.loads(r.read())
    log(f"Total profiles: {len(profiles)}")

    existing = load_existing()
    existing_lock = threading.Lock()
    log(f"Already in outreach list: {len(existing)} entries")
    log(f"Running with {WORKERS} parallel threads...\n")

    stats = {"added": 0, "existing": 0, "no_domain": 0, "no_email": 0}
    batch = []
    batch_lock = threading.Lock()

    def run(args):
        idx, p = args
        result, status = process_profile(idx, p, existing)
        if result:
            with existing_lock:
                existing.add(result["company"].lower())
            with batch_lock:
                batch.append(result)
                if len(batch) >= 10:
                    to_save = batch[:]
                    batch.clear()
                    save_to_outreach(to_save)
                    log(f"  >>> Saved 10 rows (total added so far: {stats['added'] + len(to_save)})")
        return status

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = {ex.submit(run, (i+1, p)): i for i, p in enumerate(profiles)}
        done = 0
        for f in as_completed(futures):
            status = f.result()
            stats[status if status in stats else "added"] = stats.get(status, 0) + (1 if status == "ok" else 0)
            if status == "ok":
                stats["added"] += 1
            done += 1
            if done % 50 == 0:
                log(f"\n--- Progress: {done}/{len(profiles)} processed | found: {stats['added']} ---\n")

    if batch:
        save_to_outreach(batch)

    log(f"\n=== Done ===")
    log(f"Added:        {stats['added']}")
    log(f"Already had:  {stats.get('existing', 0)}")
    log(f"No domain:    {stats.get('no_domain', 0)}")
    log(f"No email:     {stats.get('no_email', 0)}")


if __name__ == "__main__":
    main()
