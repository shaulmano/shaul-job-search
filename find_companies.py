import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook, Workbook
import os

OUTPUT_FILE = "C:\\Users\\Shaul\\Documents\\job-search\\companies.xlsx"

ALL_COMPANIES = [
    # --- Already in list (will be deduped) ---
    ("monday.com", "monday.com"),
    ("Hibob", "hibob.com"),
    ("Wix", "wix.com"),
    ("Fiverr", "fiverr.com"),
    ("JFrog", "jfrog.com"),
    ("Wiz", "wiz.io"),
    ("Gong", "gong.io"),

    # --- CYBERSECURITY ---
    ("Check Point Software", "checkpoint.com"),
    ("CyberArk", "cyberark.com"),
    ("Imperva", "imperva.com"),
    ("Radware", "radware.com"),
    ("Cybereason", "cybereason.com"),
    ("Deep Instinct", "deepinstinct.com"),
    ("Cymulate", "cymulate.com"),
    ("Cynet", "cynet.com"),
    ("Hunters", "hunters.ai"),
    ("Intezer", "intezer.com"),
    ("Cato Networks", "catonetworks.com"),
    ("Orca Security", "orca.security"),
    ("Axonius", "axonius.com"),
    ("Armis", "armis.com"),
    ("Aqua Security", "aquasec.com"),
    ("Checkmarx", "checkmarx.com"),
    ("Sternum", "sternumiot.com"),
    ("Laminar Security", "laminarsecurity.com"),
    ("Dig Security", "dig.security"),
    ("Veriti", "veriti.ai"),
    ("Sygnia", "sygnia.co"),
    ("Illusive Networks", "illusivenetworks.com"),
    ("Cybersixgill", "cybersixgill.com"),
    ("CYREBRO", "cyrebro.io"),
    ("Apiiro", "apiiro.com"),
    ("Opus Security", "opus.security"),
    ("Orca Security", "orca.security"),
    ("Noname Security", "nonamesecurity.com"),
    ("Salt Security", "salt.security"),
    ("Traceable AI", "traceable.ai"),
    ("Grip Security", "grip.security"),
    ("Dazz", "dazz.io"),
    ("Talon Cyber Security", "talon-sec.com"),
    ("Island", "island.io"),
    ("Pentera", "pentera.io"),
    ("Rezilion", "rezilion.com"),
    ("Torch Security", "torchsecurity.ai"),
    ("Gutsy", "gutsy.com"),

    # --- FINTECH ---
    ("Payoneer", "payoneer.com"),
    ("eToro", "etoro.com"),
    ("Tipalti", "tipalti.com"),
    ("Rapyd", "rapyd.com"),
    ("Global-e", "global-e.com"),
    ("Nuvei", "nuvei.com"),
    ("Melio", "meliopayments.com"),
    ("Fundbox", "fundbox.com"),
    ("Lemonade", "lemonade.com"),
    ("Next Insurance", "nextinsurance.com"),
    ("Sunbit", "sunbit.com"),
    ("Pagaya", "pagaya.com"),
    ("Earnix", "earnix.com"),
    ("Sapiens International", "sapiens.com"),
    ("Plus500", "plus500.com"),
    ("Credorax", "credorax.com"),
    ("Finaro", "finaro.com"),
    ("Mesh Payments", "meshpayments.com"),
    ("Capitolis", "capitolis.com"),
    ("FinanceKey", "financekey.com"),
    ("Entera", "entera.io"),
    ("Kramer Levin", "kramerlevin.com"),
    ("Kovrr", "kovrr.com"),
    ("Vesttoo", "vesttoo.com"),
    ("Akur8", "akur8.com"),
    ("Paysign", "paysign.com"),
    ("Covercy", "covercy.com"),
    ("Rewire", "rewire.co"),
    ("Papaya Global", "papayaglobal.com"),

    # --- AI / MACHINE LEARNING ---
    ("AI21 Labs", "ai21.com"),
    ("Tabnine", "tabnine.com"),
    ("Lightricks", "lightricks.com"),
    ("Deep Instinct", "deepinstinct.com"),
    ("BriefCam", "briefcam.com"),
    ("Aidoc", "aidoc.com"),
    ("Augury", "augury.com"),
    ("Cortica", "cortica.ai"),
    ("Deci AI", "deci.ai"),
    ("Comet", "comet.ml"),
    ("Supercom", "supercom.com"),
    ("Intelligo", "intelligo.ai"),
    ("Trigo", "trigoretail.com"),
    ("Syte", "syte.ai"),
    ("Hyro", "hyro.ai"),
    ("Quanta", "quantaco.com"),
    ("Run.ai", "run.ai"),
    ("Percepto", "percepto.ai"),
    ("Edgify", "edgify.ai"),
    ("Fabric", "fabricgroup.com"),
    ("Luminary Cloud", "luminarycloud.com"),
    ("Zencoder", "zencoder.ai"),
    ("Cohere", "cohere.com"),
    ("Glean", "glean.com"),
    ("Aleph Alpha", "aleph-alpha.com"),

    # --- HEALTHCARE / MEDTECH ---
    ("Mobileye", "mobileye.com"),
    ("OrCam", "orcam.com"),
    ("Nanox", "nanox.com"),
    ("Zebra Medical Vision", "zebra-med.com"),
    ("Sight Diagnostics", "sightdx.com"),
    ("Tyto Care", "tytocare.com"),
    ("Viz.ai", "viz.ai"),
    ("Itamar Medical", "itamarmedical.com"),
    ("Sanara MedTech", "sanaramedtech.com"),
    ("BioLineRx", "biolinerx.com"),
    ("CollPlant", "collplant.com"),
    ("MedAware", "medaware.com"),
    ("EarlySense", "earlysense.com"),
    ("Datos Health", "datos.health"),
    ("Nucleai", "nucleai.ai"),
    ("Sweetch", "sweetch.com"),
    ("Medorion", "medorion.com"),
    ("Haemonetics", "haemonetics.com"),
    ("BioHarvest Sciences", "bioharvest.com"),
    ("Todos Medical", "todosmedical.com"),
    ("Brainsway", "brainsway.com"),
    ("InVivo", "invivotherapeutics.com"),
    ("DarioHealth", "dariohealth.com"),
    ("Tali Health", "tali.health"),
    ("Binah.ai", "binah.ai"),
    ("GlucoMe", "glucome.com"),

    # --- DEVTOOLS / CLOUD ---
    ("Cloudinary", "cloudinary.com"),
    ("Swimm", "swimm.io"),
    ("Rookout", "rookout.com"),
    ("Coralogix", "coralogix.com"),
    ("Lumigo", "lumigo.io"),
    ("OverOps", "overops.com"),
    ("Sealights", "sealights.io"),
    ("Codefresh", "codefresh.io"),
    ("Snyk", "snyk.io"),
    ("Bridgecrew", "bridgecrew.io"),
    ("Epsagon", "epsagon.com"),
    ("Logz.io", "logz.io"),
    ("Frontegg", "frontegg.com"),
    ("Permit.io", "permit.io"),
    ("Elementor", "elementor.com"),
    ("WP Engine", "wpengine.com"),
    ("Novu", "novu.co"),
    ("Amplication", "amplication.com"),
    ("Blink Ops", "blinkops.com"),
    ("Firefly", "gofirefly.io"),
    ("env0", "env0.com"),
    ("Infracost", "infracost.io"),
    ("Torque", "qtorque.io"),
    ("Zesty", "zesty.co"),
    ("Cast AI", "cast.ai"),
    ("Spot by NetApp", "spot.io"),

    # --- MARKETING TECH ---
    ("Outbrain", "outbrain.com"),
    ("Taboola", "taboola.com"),
    ("Similarweb", "similarweb.com"),
    ("AppsFlyer", "appsflyer.com"),
    ("IronSource", "ironsource.com"),
    ("Skai", "skai.com"),
    ("Optimove", "optimove.com"),
    ("Perion", "perion.com"),
    ("DoubleVerify", "doubleverify.com"),
    ("Kaltura", "kaltura.com"),
    ("Bizzabo", "bizzabo.com"),
    ("Oktopost", "oktopost.com"),
    ("Lusha", "lusha.com"),
    ("Demandbase", "demandbase.com"),
    ("Seismic", "seismic.com"),
    ("Yotpo", "yotpo.com"),
    ("Feedvisor", "feedvisor.com"),
    ("Namogoo", "namogoo.com"),
    ("Nosto", "nosto.com"),
    ("Glassbox", "glassbox.com"),
    ("Contentsquare", "contentsquare.com"),
    ("Unbounce", "unbounce.com"),

    # --- HR TECH ---
    ("Comeet", "comeet.com"),
    ("Talenya", "talenya.com"),
    ("Empact", "empact.io"),
    ("HiBob", "hibob.com"),
    ("Guesty", "guesty.com"),
    ("Jobvite", "jobvite.com"),
    ("Sparrow", "trysparrow.com"),
    ("GrowthSpace", "growthspace.com"),
    ("Riseup", "riseup.ai"),
    ("Pendo", "pendo.io"),
    ("SHL", "shl.com"),
    ("Sapling", "saplinghr.com"),
    ("Leapsome", "leapsome.com"),
    ("Deel", "letsdeel.com"),
    ("Connecteam", "connecteam.com"),
    ("Workiz", "workiz.com"),

    # --- ENTERPRISE SOFTWARE ---
    ("NICE", "nice.com"),
    ("Amdocs", "amdocs.com"),
    ("Comverse", "comverse.com"),
    ("ClickSoftware", "clicksoftware.com"),
    ("Allot", "allot.com"),
    ("Varonis", "varonis.com"),
    ("Cellebrite", "cellebrite.com"),
    ("RADCOM", "radcom.com"),
    ("Silicom", "silicom-usa.com"),
    ("AudioCodes", "audiocodes.com"),
    ("Ceragon Networks", "ceragon.com"),
    ("Commvault", "commvault.com"),
    ("LivePerson", "liveperson.com"),
    ("SysAid", "sysaid.com"),
    ("Atera", "atera.com"),
    ("WalkMe", "walkme.com"),
    ("Sisense", "sisense.com"),
    ("GreenRoad", "greenroad.com"),
    ("ClickTale", "clicktale.com"),
    ("Pyramid Analytics", "pyramidanalytics.com"),
    ("ECI Telecom", "ecitele.com"),
    ("NICE inContact", "niceincontact.com"),
    ("LiveU", "liveu.tv"),
    ("Allot", "allot.com"),
    ("Synamedia", "synamedia.com"),
    ("Verint", "verint.com"),
    ("NICE Actimize", "niceactimize.com"),

    # --- GAMING ---
    ("Playtika", "playtika.com"),
    ("Moon Active", "moonactive.com"),
    ("Plarium", "plarium.com"),
    ("Overwolf", "overwolf.com"),
    ("Solitaire Grand Harvest", "supertreat.com"),
    ("Innplay Labs", "innplaylabs.com"),
    ("Jelly Button Games", "jellybuttongames.com"),
    ("Sciplay", "sciplay.com"),
    ("Bit Reactor", "bitreactor.io"),

    # --- AUTONOMOUS / TRANSPORTATION ---
    ("Via", "ridewithvia.com"),
    ("Innoviz", "innoviz.tech"),
    ("Arbe Robotics", "arberobotics.com"),
    ("Brodmann17", "brodmann17.com"),
    ("Nexar", "nexar.com"),
    ("Tactile Mobility", "tactilemobility.com"),
    ("Cognata", "cognata.com"),
    ("Foretellix", "foretellix.com"),
    ("TriEye", "trieye.tech"),
    ("Foresight", "foresightauto.com"),

    # --- AGTECH ---
    ("Taranis", "taranis.ag"),
    ("Phytech", "phytech.com"),
    ("CropX", "cropx.com"),
    ("Prospera", "prospera.ag"),
    ("BeeHero", "beehero.io"),
    ("SupPlant", "supplant.com"),
    ("Arable", "arable.com"),
    ("Netafim", "netafim.com"),

    # --- SEMICONDUCTORS ---
    ("Tower Semiconductor", "towersemi.com"),
    ("Camtek", "camtek.com"),
    ("CEVA", "ceva-dsp.com"),
    ("Valens Semiconductor", "valens-semi.com"),
    ("DSP Group", "dspg.com"),
    ("Silicom", "silicom-usa.com"),
    ("Hailo", "hailo.ai"),
    ("Kneron", "kneron.com"),
    ("Greenfield Acquisition", "greenfieldacquisition.com"),

    # --- ENERGY TECH ---
    ("SolarEdge", "solaredge.com"),
    ("Ormat Technologies", "ormat.com"),
    ("Energix", "energix.co.il"),
    ("Brenmiller Energy", "bren-energy.com"),
    ("StoreDot", "store-dot.com"),
    ("Irp Nexus", "irpnexus.com"),

    # --- E-COMMERCE / RETAIL TECH ---
    ("Riskified", "riskified.com"),
    ("Global-e", "global-e.com"),
    ("Yotpo", "yotpo.com"),
    ("Syte", "syte.ai"),
    ("Namogoo", "namogoo.com"),
    ("Feedvisor", "feedvisor.com"),
    ("Compari Tech", "comparitech.com"),
    ("Bazaarvoice", "bazaarvoice.com"),
    ("Nuvei", "nuvei.com"),

    # --- LEGAL TECH ---
    ("LawGeex", "lawgeex.com"),
    ("Darrow", "darrow.ai"),
    ("Lexion", "lexion.ai"),
    ("Evisort", "evisort.com"),

    # --- REAL ESTATE TECH ---
    ("HonestBuildings", "honestbuildings.com"),
    ("Entera", "entera.io"),
    ("Guesty", "guesty.com"),
    ("Matterport", "matterport.com"),
    ("Skyline AI", "skyline.ai"),
    ("Building Engines", "buildingengines.com"),

    # --- EDTECH ---
    ("Typesense", "typesense.org"),
    ("Fiverr", "fiverr.com"),
    ("Kahoot", "kahoot.com"),
    ("Comigo", "comigo.ai"),
    ("Cognii", "cognii.com"),
    ("Sela", "sela.co.il"),

    # --- COMMUNICATION / COLLAB ---
    ("Loom", "loom.com"),
    ("Blink", "joinblink.com"),
    ("Demostack", "demostack.com"),
    ("Strigo", "strigo.io"),
    ("LivePerson", "liveperson.com"),
    ("Vonage", "vonage.com"),
    ("Amdocs", "amdocs.com"),
    ("Kaltura", "kaltura.com"),
    ("Riversideo FM", "riverside.fm"),
    ("Miro", "miro.com"),

    # --- OPERATIONS / FIELD SERVICE ---
    ("Riskified", "riskified.com"),
    ("Bringg", "bringg.com"),
    ("Connecteam", "connecteam.com"),
    ("Workiz", "workiz.com"),
    ("Optibus", "optibus.com"),
    ("Routemaster", "routemaster.app"),
    ("FreshDesk", "freshdesk.com"),
    ("Kustomer", "kustomer.com"),
    ("Zendesk", "zendesk.com"),
    ("Totango", "totango.com"),
    ("Gainsight", "gainsight.com"),
    ("Salto", "salto.io"),
]


def dedup(companies):
    seen_domains = set()
    result = []
    for name, domain in companies:
        if domain.lower() not in seen_domains:
            seen_domains.add(domain.lower())
            result.append((name, domain))
    return result


def load_existing_domains():
    if not os.path.exists(OUTPUT_FILE):
        return set()
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    domains = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            domains.add(str(row[1]).strip().lower())
    return domains


def save_to_excel():
    companies = dedup(ALL_COMPANIES)
    existing = load_existing_domains()

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Companies"
        ws.append(["Company Name", "Domain"])
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30

    added = 0
    for name, domain in companies:
        if domain.lower() not in existing:
            ws.append([name, domain])
            existing.add(domain.lower())
            added += 1

    wb.save(OUTPUT_FILE)
    print(f"Added {added} new companies.")
    print(f"Total in file: {ws.max_row - 1}")


if __name__ == "__main__":
    save_to_excel()
