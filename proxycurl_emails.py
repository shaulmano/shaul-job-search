import sys
sys.stdout.reconfigure(encoding='utf-8')

import csv
import time
import requests
from openpyxl import load_workbook

PROXYCURL_API_KEY = "GI2eUSb83NIBWeM-u-DKpw"

RECRUITERS_FILE = r"C:\Users\Shaul\Documents\job-search\phantom_recruiters.csv"
OUTREACH_FILE   = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"
PROGRESS_FILE   = r"C:\Users\Shaul\Documents\job-search\proxycurl_done.txt"


def load_done():
    try:
        with open(PROGRESS_FILE, encoding='utf-8') as f:
            return set(line.strip() for line in f if line.strip())
    except FileNotFoundError:
        return set()


def mark_done(url):
    with open(PROGRESS_FILE, 'a', encoding='utf-8') as f:
        f.write(url + '\n')


def load_existing_emails():
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active
    return set(
        str(row[3].value or '').strip().lower()
        for row in ws.iter_rows(min_row=2)
        if row[3].value
    )


def add_to_outreach(first, last, company, email, linkedin_url):
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active
    ws.append([first, last, company, email, linkedin_url, 'Pending', ''])
    wb.save(OUTREACH_FILE)


def get_email(linkedin_url, api_key):
    headers = {'Authorization': f'Bearer {api_key}'}
    params  = {
        'linkedin_profile_url': linkedin_url,
        'personal_email': 'include',
    }
    resp = requests.get(
        'https://nubela.co/proxycurl/api/v2/linkedin',
        headers=headers,
        params=params,
        timeout=20
    )
    if resp.status_code == 200:
        data = resp.json()
        emails = data.get('personal_emails') or []
        if not emails:
            # Try work email too
            emails = data.get('work_email') and [data['work_email']] or []
        company = ''
        exp = data.get('experiences') or []
        if exp:
            company = exp[0].get('company', '')
        return emails[0] if emails else None, company
    return None, ''


def main():
    if not PROXYCURL_API_KEY:
        print("שגיאה: הכנס API Key ב-PROXYCURL_API_KEY בסקריפט.")
        return

    with open(RECRUITERS_FILE, encoding='utf-8-sig') as f:
        recruiters = list(csv.DictReader(f))

    done         = load_done()
    existing     = load_existing_emails()
    pending      = [r for r in recruiters if r['linkedinProfileUrl'].strip() not in done]

    print(f"סה\"כ פרופילים  : {len(recruiters)}")
    print(f"כבר עובדו      : {len(done)}")
    print(f"נותרו לעיבוד   : {len(pending)}")
    print()

    found = 0
    skipped = 0

    for i, r in enumerate(pending, 1):
        url   = r['linkedinProfileUrl'].strip()
        first = r['firstName'].strip()
        last  = r['lastName'].strip()

        print(f"[{i}/{len(pending)}] {first} {last} ... ", end='', flush=True)

        email, company = get_email(url, PROXYCURL_API_KEY)
        mark_done(url)

        if email and email.lower() not in existing:
            add_to_outreach(first, last, company, email, url)
            existing.add(email.lower())
            found += 1
            print(f"נמצא: {email}")
        else:
            skipped += 1
            print("אין מייל ציבורי")

        time.sleep(1)  # מניעת rate limit

    print(f"\n=== סיום ===")
    print(f"מיילים חדשים נוספו : {found}")
    print(f"ללא מייל ציבורי    : {skipped}")


if __name__ == '__main__':
    main()
