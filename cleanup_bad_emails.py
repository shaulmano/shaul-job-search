import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
import re

OUTREACH_FILE = r"C:\Users\Shaul\Documents\job-search\outreach_list.xlsx"

# Domains known to be wrong — never a real company email
BAD_EMAIL_DOMAINS = {
    "youtube.com", "twitter.com", "instagram.com", "tiktok.com",
    "rocketreach.co", "rocketreach.io", "contactout.com", "finalscout.com",
    "leadiq.com", "nytimes.com", "timesofisrael.com", "jpost.com",
    "medium.com", "reddit.com", "quora.com", "slideshare.net",
    "investing.com", "theorg.com", "growjo.com", "beststartup.asia",
    "kycisrael.com", "linktr.ee", "affplus.com", "volza.com", "dnb.com",
    "israelforever.org", "israelgives.org", "calcalistech.com",
    "techrepublic.com", "talkbusiness.net", "drushim.co.il",
    "alljobs.co.il", "github.com", "hbomax.com", "palestinechronicle.com",
    "cufi.org", "ritholtz.com", "gigaom.com", "ramp.directory",
    "mastersportal.com", "webatla.com", "intch.org", "ebenezer.org.il",
    "timetoast.com", "noatishby.com", "nortech-platform.com",
}

wb = load_workbook(OUTREACH_FILE)
ws = wb.active

rows_to_delete = []
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    email = str(row[3].value or "").strip().lower()
    status = str(row[5].value or "").strip()
    if status == "Pending" and email:
        domain = email.split("@")[-1] if "@" in email else ""
        if domain in BAD_EMAIL_DOMAINS:
            rows_to_delete.append(i)
            print(f"  Delete row {i}: {email}")

# Delete in reverse order to preserve row numbers
for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)

wb.save(OUTREACH_FILE)
print(f"\nDeleted {len(rows_to_delete)} bad rows.")
