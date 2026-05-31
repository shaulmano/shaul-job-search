import sys
sys.stdout.reconfigure(encoding='utf-8')

import smtplib, os, time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
from openpyxl import load_workbook
from config import (
    GMAIL_ADDRESS, GMAIL_APP_PASSWORD, SENDER_NAME,
    SENDER_PHONE, SENDER_LINKEDIN, PDF_ATTACHMENT, CV_PM, CV_QA, OUTREACH_FILE
)

DAILY_LIMIT   = 50
BOUNCE_WAIT   = 3600  # 1 hour in seconds
LOG_FILE      = os.path.join(os.path.dirname(__file__), 'daily_outreach.log')

EMAIL_SUBJECT = "25 Years in QA & Project Manager - Worth a 15-Minute Call?"
EMAIL_BODY = """\
Hi {first_name},

I'm reaching out because I believe my background could be a strong fit for {company}.

I'm Shaul Mano — a Senior Program Manager and QA Leader with 25+ years of experience at companies like RSA and Symantec.

If there's anything relevant open at {company}, I'd love a quick 15-minute call.
I've attached two CVs - one focused on Program Management and one on QA Leadership - so you can share whichever fits best.

Best,
Shaul Mano
{phone}
{linkedin}
"""


def log(msg):
    line = f"[{date.today()} {time.strftime('%H:%M:%S')}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def build_email(to_address, first_name, company):
    msg = MIMEMultipart()
    msg["From"] = f"{SENDER_NAME} <{GMAIL_ADDRESS}>"
    msg["To"]   = to_address
    msg["Subject"] = EMAIL_SUBJECT
    body = EMAIL_BODY.format(
        first_name=first_name, company=company,
        phone=SENDER_PHONE, linkedin=SENDER_LINKEDIN,
    )
    msg.attach(MIMEText(body, "plain"))
    for attachment in [PDF_ATTACHMENT, CV_PM, CV_QA]:
        if os.path.exists(attachment):
            with open(attachment, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",
                            f'attachment; filename="{os.path.basename(attachment)}"')
            msg.attach(part)
    return msg


def send_batch():
    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    pending = [
        {"row_num": i, "first": str(row[0].value or "").strip(),
         "last": str(row[1].value or "").strip(),
         "company": str(row[2].value or "").strip(),
         "email": str(row[3].value or "").strip()}
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2)
        if row[5].value == "Pending" and str(row[3].value or "").strip()
    ]

    if not pending:
        log("No pending emails — done.")
        return 0

    batch = pending[:DAILY_LIMIT]
    log(f"Sending {len(batch)} emails ({len(pending)} pending total)...")

    sent = 0
    for item in batch:
        try:
            msg = build_email(item["email"], item["first"], item["company"])
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
                server.sendmail(GMAIL_ADDRESS, item["email"], msg.as_string())
            ws.cell(row=item["row_num"], column=6).value = "Sent"
            ws.cell(row=item["row_num"], column=7).value = str(date.today())
            wb.save(OUTREACH_FILE)
            sent += 1
            log(f"  [{sent}/{len(batch)}] {item['first']} {item['last']} <{item['email']}>")
            time.sleep(3)
        except Exception as e:
            log(f"  ERROR {item['email']}: {e}")

    log(f"Sent {sent}/{len(batch)} emails.")
    return sent


def check_bounces():
    import importlib.util, pathlib
    script = pathlib.Path(__file__).parent / "check_bounces.py"
    spec = importlib.util.spec_from_file_location("check_bounces", script)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)


if __name__ == "__main__":
    log("=== Daily outreach started ===")
    sent = send_batch()

    if sent > 0:
        log(f"Waiting {BOUNCE_WAIT // 60} minutes before checking bounces...")
        time.sleep(BOUNCE_WAIT)

    log("Checking bounces...")
    try:
        check_bounces()
    except Exception as e:
        log(f"Bounce check error: {e}")

    log("=== Daily outreach done ===")
