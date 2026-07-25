"""
Adds a LinkedIn column to vc_outreach_list.xlsx and fills in the company
pages I could verify. Patches the live file in place, so Sent/Bounced
statuses survive. Safe to re-run.

Only funds whose LinkedIn page I actually confirmed are here. The rest are
left blank on purpose - a wrong LinkedIn URL wastes your time the same way
a wrong email address does.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

VC_FILE = r"C:\Users\Shaul\Documents\job-search\vc_outreach_list.xlsx"

LINKEDIN = {
    "Viola Ventures": "https://il.linkedin.com/company/violaventures",
    "F2 Venture Capital": "https://il.linkedin.com/company/f2-capital",
    "Hanaco Ventures": "https://www.linkedin.com/company/hanacoventures/",
    "Grove Ventures": "https://www.linkedin.com/company/grove-ventures",
    "Vertex Ventures Israel": "https://il.linkedin.com/company/vertex-ventures-israel",
    "Aleph": "https://il.linkedin.com/company/aleph-vc",
    "Cyberstarts": "https://www.linkedin.com/company/cyberstarts",
    "Greenfield Partners": "https://www.linkedin.com/company/greenfieldlp",
    "Qumra Capital": "https://www.linkedin.com/company/qumra-capital",
    "Entree Capital": "https://www.linkedin.com/company/entree-capital",
    "State of Mind Ventures": "https://www.linkedin.com/company/state-of-mind-ventures",
    "Magma Venture Partners": "https://www.linkedin.com/company/magma-venture-partners",
    "Hyperwise Ventures": "https://il.linkedin.com/company/hyperwise-ventures",
}


def main():
    wb = load_workbook(VC_FILE)
    ws = wb.active
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}

    if "LinkedIn" not in headers:
        col = ws.max_column + 1
        ws.cell(row=1, column=col).value = "LinkedIn"
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=1).font.copy()
        ws.cell(row=1, column=col).fill = ws.cell(row=1, column=1).fill.copy()
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 52
        headers["LinkedIn"] = col - 1
        print("added LinkedIn column")

    filled = 0
    for row in ws.iter_rows(min_row=2):
        fund = str(row[headers["Fund"]].value or "").strip()
        if fund in LINKEDIN:
            row[headers["LinkedIn"]].value = LINKEDIN[fund]
            filled += 1

    wb.save(VC_FILE)
    print(f"filled {filled} LinkedIn URLs")
    print(f"saved {VC_FILE}")


if __name__ == "__main__":
    main()
