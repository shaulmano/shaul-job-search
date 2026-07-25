import sys
sys.stdout.reconfigure(encoding='utf-8')

import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
from openpyxl import load_workbook
from config import GMAIL_ADDRESS, GMAIL_APP_PASSWORD, SENDER_NAME, SENDER_PHONE, SENDER_LINKEDIN

DEXHIVE_FILE = r"C:\Users\Shaul\Documents\מגייסות_דקסהייב.xlsx"
CV_PATH      = r"C:\Users\Shaul\Documents\job-search\Shaul_Mano_2026.pdf"
DAILY_LIMIT  = 200

EMAIL_SUBJECT = "25 Years in QA & Project Manager - Worth a 15-Minute Call?"

EMAIL_BODY_TEMPLATE = """\
Hi {english_name},

I'm Shaul Mano — a Senior Program Manager and QA Leader with 25+ years of experience at companies like RSA and Symantec.

If there's anything relevant open in your pipeline, I'd love a quick 15-minute call.
I've attached my CV for your review.

Best,
Shaul Mano
""" + SENDER_PHONE + "\n" + SENDER_LINKEDIN + "\n"


def build_email(to_address, english_name):
    msg = MIMEMultipart()
    msg["From"] = f"{SENDER_NAME} <{GMAIL_ADDRESS}>"
    msg["To"]   = to_address
    msg["Subject"] = EMAIL_SUBJECT
    body = EMAIL_BODY_TEMPLATE.format(english_name=english_name or "there")
    msg.attach(MIMEText(body, "plain"))

    if os.path.exists(CV_PATH):
        with open(CV_PATH, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(CV_PATH)}"')
        msg.attach(part)
    else:
        print(f"  WARNING: CV not found: {CV_PATH}")

    return msg


def send_email(msg, to_address):
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, to_address, msg.as_string())


def main():
    wb = load_workbook(DEXHIVE_FILE)
    ws = wb.active

    # Ensure Status and Date columns exist
    if ws.cell(1, 4).value is None:
        ws.cell(1, 4).value = "סטטוס"
    if ws.cell(1, 5).value is None:
        ws.cell(1, 5).value = "תאריך שליחה"

    # Collect pending rows (no status or status != "Sent")
    pending = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        english  = str(row[0].value or "").strip().strip("()")
        email    = str(row[1].value or "").strip()
        status   = str(row[2].value or "").strip() if len(row) > 2 else ""

        if not email or status == "Sent":
            continue

        pending.append({
            "row_num": i,
            "english": english,
            "email":   email,
        })

    if not pending:
        print("אין כתובות ממתינות לשליחה.")
        return

    print(f"נמצאו {len(pending)} כתובות ממתינות.")
    batch = pending[:DAILY_LIMIT]

    # Preview
    print("\n" + "=" * 60)
    print(f"תצוגה מקדימה — 3 ראשונות מתוך {len(batch)}:")
    print("=" * 60)
    for item in batch[:3]:
        print(f"\nאל: {item['email']}  |  {item['english']}")
        print("-" * 60)
        print(EMAIL_BODY_TEMPLATE.format(english_name=item['english'] or "there"))
        print(f"קובץ מצורף: {os.path.basename(CV_PATH)}")
        print("=" * 60)

    print(f"\nשולח {len(batch)} מיילים...\n")
    sent_count = 0
    for item in batch:
        try:
            msg = build_email(item["email"], item["english"])
            send_email(msg, item["email"])
            ws.cell(row=item["row_num"], column=3).value = "Sent"
            ws.cell(row=item["row_num"], column=4).value = str(date.today())
            wb.save(DEXHIVE_FILE)
            sent_count += 1
            print(f"  [{sent_count}/{len(batch)}] {item['email']}")
        except Exception as e:
            print(f"  ERROR → {item['email']}: {e}")

    print(f"\nסה\"כ נשלחו: {sent_count} / {len(batch)}")


if __name__ == "__main__":
    main()
