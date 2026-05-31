import sys
sys.stdout.reconfigure(encoding='utf-8')

import requests
import time
from openpyxl import load_workbook
from config import OUTREACH_FILE

REOON_API_KEY = "nPHnDjoPAwb6BT2HiRdPhrYOS33R2ooi"

def verify_email(email):
    try:
        resp = requests.get(
            "https://emailverifier.reoon.com/api/v1/verify",
            params={"email": email, "key": REOON_API_KEY, "mode": "quick"},
            timeout=15
        )
        data = resp.json()
        status = data.get("status", "unknown")
        return status
    except Exception as e:
        print(f"  Error checking {email}: {e}")
        return "unknown"

def main():
    if not REOON_API_KEY:
        print("ERROR: הכנס REOON_API_KEY בסקריפט")
        return

    wb = load_workbook(OUTREACH_FILE)
    ws = wb.active

    pending = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status = str(row[5].value or "").strip()
        email  = str(row[3].value or "").strip()
        if status == "Pending" and email:
            pending.append((i, email))

    print(f"Pending emails to verify: {len(pending)}\n")

    removed = 0
    for i, (row_idx, email) in enumerate(pending, 1):
        status = verify_email(email)
        print(f"[{i}/{len(pending)}] {email} → {status}")

        if status in ("invalid", "disposable", "spamtrap"):
            for row in ws.iter_rows(min_row=row_idx, max_row=row_idx):
                row[5].value = "Bounced"
            removed += 1
            print(f"  → marked as Bounced")

        time.sleep(0.3)

    wb.save(OUTREACH_FILE)
    print(f"\nDone. Removed {removed} invalid emails from Pending.")

if __name__ == "__main__":
    main()
